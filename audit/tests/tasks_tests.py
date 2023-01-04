import tempfile
import zipfile

import openpyxl
import pytest
from django.core import serializers
from django.core.files import File

from audit.models import Audit, AuditFirm, Auditor, AuditStatus
from audit.tasks import export_audit
from fieldwork.constants import ER_STATUS_DICT, REQ_STATUS_DICT, TEST_RESULTS
from fieldwork.models import Criteria, Evidence, Requirement, Test
from user.constants import AUDITOR_ADMIN
from user.tests import create_user_auditor

from .factory import create_audit


@pytest.fixture
def audit(graphql_organization, graphql_audit_firm):
    return create_audit(
        organization=graphql_organization,
        name='Laika Dev Soc 2 Type 1 Audit 2022',
        audit_firm=graphql_audit_firm,
    )


@pytest.fixture
def criteria_list():
    return [
        Criteria.objects.create(display_id='CC1.1', description='yyy'),
        Criteria.objects.create(display_id='CC1.2', description='weersdasd'),
    ]


@pytest.fixture
def evidence_opened(audit):
    return Evidence.objects.create(
        audit=audit,
        display_id='ER-1',
        name='ER1',
        instructions='yyyy',
        status=ER_STATUS_DICT['Open'],
    )


@pytest.fixture
def evidence_accepted(audit):
    return Evidence.objects.create(
        audit=audit,
        display_id='ER-2',
        name='ER2',
        instructions='aaaaaaa',
        status=ER_STATUS_DICT['Auditor Accepted'],
    )


@pytest.fixture
def accepted_er_attachments(evidence_accepted):
    for i in range(3):
        file_name = f'Attachment{i}.pdf'
        evidence_accepted.add_attachment(
            file_name=file_name,
            file=File(name=file_name, file=tempfile.TemporaryFile()),
        )


@pytest.fixture
def requirements(audit):
    requirements = []
    for i in range(5):
        # To test only non excluded and non deleted requirements
        # are linked
        excluded_prop = (
            {'is_deleted': True, 'exclude_in_report': True} if i == 0 else {}
        )
        requirement = Requirement.objects.create(
            audit=audit,
            display_id=f'LCL-{i}',
            name=f'LCL-{i}',
            status=REQ_STATUS_DICT['Completed'],
            **excluded_prop,
        )
        requirements.append(requirement)

    return requirements


checklist = '''<html><head></head><body><p>blabla</p></body></html>'''


@pytest.fixture
def tests(audit, requirements):
    tests = [
        Test.objects.create(
            display_id='Test-1',
            name='Test1',
            result='exceptions_noted',
            checklist=checklist,
            requirement=requirements[0],
        ),
        Test.objects.create(
            display_id='Test-2',
            name='Test2',
            result='exceptions_noted',
            checklist=checklist,
            notes='The note',
            requirement=requirements[1],
        ),
        Test.objects.create(
            display_id='Test-3',
            name='Test3',
            result='exceptions_noted',
            checklist=checklist,
            notes='The note',
            requirement=requirements[1],
        ),
        Test.objects.create(
            display_id='Test-4',
            name='Test4',
            result='exceptions_noted',
            checklist=checklist,
            notes='The note',
            requirement=requirements[1],
        ),
    ]

    return tests


@pytest.fixture
def auditor_user(graphql_audit_firm: AuditFirm) -> Auditor:
    return create_user_auditor(
        email='johndoe@heylaika.com',
        role=AUDITOR_ADMIN,
        with_audit_firm=True,
        audit_firm=graphql_audit_firm.name,
    )


def validate_requirements(workbook):
    req_ids = ['LCL-1', 'LCL-2', 'LCL-3', 'LCL-4', 'LCL-5']
    req_accepted_ers = 'ER-2'
    req_tests = 'Test-2,Test-3,Test-4'

    req_sheet = workbook['requirements']
    for row in req_sheet.iter_rows(min_row=2, max_col=7, max_row=4):
        for cell in row:
            assert (
                cell.value in req_ids
                or cell.value == req_accepted_ers
                or cell.value == req_tests
                or cell.value is None
            )  # For empty testers/reviewers


def validate_criteria(workbook):
    criteria_ids = ['CC1.1', 'CC1.2']
    criteria_desc = ['yyy', 'weersdasd']
    criteria_reqs = 'LCL-1,LCL-2,LCL-3,LCL-4'

    criteria_sheet = workbook['criteria']
    for row in criteria_sheet.iter_rows(min_row=2, max_col=3, max_row=2):
        for cell in row:
            assert (
                cell.value in criteria_ids
                or cell.value in criteria_desc
                or cell.value == criteria_reqs
            )


def validate_tests(workbook):
    test_ids = ['Test-2', 'Test-3', 'Test-4']
    test_names = ['Test2', 'Test3', 'Test4']
    test_req = 'LCL-1'
    notes = 'The note'
    test_result = dict(TEST_RESULTS)['exceptions_noted']
    plain_checklist = 'blabla'
    test_sheet = workbook['tests']
    for row in test_sheet.iter_rows(min_row=2, max_col=6, max_row=2):
        for cell in row:
            assert (
                cell.value in test_ids
                or cell.value in test_names
                or cell.value == test_req
                or cell.value == notes
                or cell.value == test_result
                or cell.value == plain_checklist
            )


def validate_evidence_requests(workbook):
    er_ids = ['ER-2']
    er_names = ['ER2']
    er_reqs = 'LCL-1,LCL-2,LCL-3,LCL-4'
    er_sheet = workbook['evidence_request']
    for row in er_sheet.iter_rows(min_row=2, max_col=2, max_row=1):
        for cell in row:
            assert (
                cell.value in er_ids or cell.value in er_names or cell.value == er_reqs
            )


def validate_attachments(zip_folder):
    for file in zip_folder.namelist():
        if file.startswith('ER-2/'):
            assert (
                file == 'ER-2/Attachment0.pdf'
                or file == 'ER-2/Attachment1.pdf'
                or file == 'ER-2/Attachment2.pdf'
            )


@pytest.mark.functional
def test_export_audit_file(
    audit,
    requirements,
    criteria_list,
    tests,
    evidence_accepted,
    evidence_opened,
    accepted_er_attachments,
    auditor_user,
):
    AuditStatus.objects.create(
        audit=audit,
        completed=True,
    )
    for requirement in requirements:
        requirement.criteria.set(criteria_list)
        requirement.evidence.set([evidence_opened, evidence_accepted])

    export_audit.delay(audit.id, serializers.serialize('json', [auditor_user.user])[0])

    updated_audit = Audit.objects.get(id=audit.id)

    exported_file = updated_audit.exported_audit_file
    file_name = '_SOC_2_TYPE_1_Audit.xlsx'

    with zipfile.ZipFile(exported_file) as zf:
        with zf.open(file_name, mode='r') as audit_file:
            wb = openpyxl.load_workbook(audit_file, read_only=True)
            validate_criteria(wb)
            validate_requirements(wb)
            validate_tests(wb)
            validate_evidence_requests(wb)
            validate_attachments(zf)

from django.core.files import File
from openpyxl import load_workbook

from laika.utils.exceptions import ServiceException


class MonitorExcelReader:
    def __init__(self, excel_file: File) -> None:
        self._excel_file = excel_file
        self._is_excel_file()
        self._load_workbook_from_file()

    def _is_excel_file(self):
        if not self._excel_file.name.endswith('.xlsx'):
            raise ServiceException('Monitor excel file must be .xlsx')

    def _load_workbook_from_file(self):
        self._workbook = load_workbook(self._excel_file.file)

    def _is_sheet_in_workbook(self, sheet_name: str):
        return sheet_name in self._workbook.sheetnames

    def count_items_on_sheet(self, sheet_name: str):
        if self._is_sheet_in_workbook(sheet_name):
            sheet = self._workbook[sheet_name]
            rows = 0
            data_initial_row = 2
            for row in sheet.iter_rows(min_row=data_initial_row):
                if not all(cell.value is None for cell in row):
                    rows += 1
            return rows
        else:
            return 0

import json
import logging
from datetime import datetime
from multiprocessing.pool import ThreadPool
from typing import List, Optional, Tuple
from zipfile import ZipFile

import graphene
import reversion
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Prefetch, Q, QuerySet
from django.utils import timezone
from openpyxl import load_workbook

import evidence.constants as constants
from access_review.mutations import update_access_review_from_action_item
from action_item.models import ActionItem, ActionItemStatus
from alert.constants import ALERT_TYPES
from audit.models import Audit
from certification.models import (
    ArchivedUnlockedOrganizationCertification,
    Certification,
    UnlockedOrganizationCertification,
)
from control.constants import (
    CAPITALIZE_YES,
    CONTROL_TYPE,
    CUSTOM_PREFIX,
    DONE,
    FAILED,
    IN_PROGRESS,
    LAI_REF_ID,
    MAPPING_PROFILE_NAME,
    MAPPING_SHEET,
    MAX_OWNER_LIMIT_PER_CONTROL,
    MIGRATION_ID,
    MY_COMPLIANCE_MIGRATION_FLAGS,
    REQUIRED_EVIDENCE_NO,
    REQUIRED_EVIDENCE_YES,
    SEED_PROFILE_NAME,
    SEED_PROFILES_MAPPING,
    STATUS,
    MetadataFields,
)
from control.inputs import (
    AddControlActionItemInput,
    ControlEvidenceInput,
    CreateControlInput,
    DeleteControlEvidenceInput,
    DeleteControlsInput,
    UpdateControlActionItemInput,
    UpdateControlActionItemsInput,
    UpdateControlFamilyOwnerInput,
    UpdateControlInput,
)
from control.models import Control, ControlComment, ControlTag
from control.types import ControlActionItemType, ControlType
from control.utils.status import can_transition_status
from evidence.models import Evidence
from feature.constants import playbooks_feature_flag
from feature.helpers import create_or_enable_flag
from feature.models import Flag
from laika.auth import login_required, permission_required
from laika.backends.concierge_backend import ConciergeAuthenticationBackend
from laika.backends.laika_backend import AuthenticationBackend
from laika.decorators import concierge_service, laika_service, service
from laika.utils.exceptions import (
    GENERIC_FILES_ERROR_MSG,
    ServiceException,
    service_exception,
)
from laika.utils.get_organization_by_user_type import get_organization_by_user_type
from laika.utils.history import create_revision
from laika.utils.websocket import send_ws_message_to_group
from organization.inputs import MigrateOrganizationPayload
from organization.models import Organization
from program.models import SubTask
from seeder.constants import CX_APP_ROOM
from seeder.models import MyComplianceMigration, Seed, SeedProfile
from seeder.seeders.commons import get_headers
from tag.models import Tag
from user.helpers import get_user_by_email
from user.inputs import DeleteInput
from user.models import User

from .evidence_handler import (
    add_control_documents_or_laika_papers,
    add_control_officers,
    add_control_other_evidence,
    add_control_policy,
    add_control_teams,
    add_note_to_control,
    delete_evidence,
    upload_control_file,
    validate_documents_already_exist,
)
from .helpers import (
    NotSubtaskFoundInMappingFile,
    annotate_action_items_control_count,
    bulk_update_control_action_items_due_date,
    bulk_update_control_action_items_owner,
    set_assignees_to_all_child_action_items,
)

logger = logging.getLogger('control_mutations')
pool = ThreadPool()


def strip_comment_reply_content(content):
    return content.strip() if content else None


def validate_comment_content(content, user_email):
    if not content:
        logger.error(f'Trying to create a comment with empty contentuser {user_email}')
        raise ServiceException('Comment content cannot be empty')


@transaction.atomic
def tag_control(organization_id, control, tags):
    safer_tags = tags or []
    tags_objs = []
    for t in safer_tags:
        populate_tags_to_create(organization_id, control, t, tags_objs)
    ControlTag.objects.bulk_create(objs=tags_objs)


@transaction.atomic
def update_tag_control(organization_id, control, tags):
    safer_tags = tags or []

    tag_dict = dict()
    for tag in Tag.objects.filter(name__in=safer_tags):
        tag_dict[tag.name] = tag.id

    # delete the tags removed by the user.
    ControlTag.objects.filter(control_id=control).exclude(
        tag__in=list(tag_dict.values()), control=control
    ).delete()

    tags_to_update = []
    for t in safer_tags:
        # create only new tags.
        if (
            tag_dict
            and not ControlTag.objects.filter(
                tag=tag_dict.get(t), control=control
            ).exists()
        ):
            populate_tags_to_create(organization_id, control, t, tags_to_update)

    ControlTag.objects.bulk_create(objs=tags_to_update)


def populate_tags_to_create(organization_id, control, t, tags_to_update):
    tag, _ = Tag.objects.get_or_create(name=t.strip(), organization_id=organization_id)
    tags_to_update.append(ControlTag(tag=tag, control=control))


def validate_owners_quantity(owners):
    if len(owners) > MAX_OWNER_LIMIT_PER_CONTROL:
        message = f'Maximum owners allowed is: {MAX_OWNER_LIMIT_PER_CONTROL}'
        raise ServiceException(message)


def validate_owners_empty_values(owners_email):
    if '' in owners_email:
        raise ServiceException('The owner email can not be an empty string')


def get_control_comment(input):
    control_comment = ControlComment.objects.get(
        control_id=input.control_id, comment_id=input.comment_id
    )
    return control_comment


def send_comment_email_alert(alert, control_related):
    alert.send_comment_control_alert_email(control_related=control_related)


def create_mention_alerts(mentions, alert_type=ALERT_TYPES['CONTROL_MENTION']):
    for mention in mentions:
        room_id = mention.user.organization.id
        alert = mention.create_mention_alert(room_id, alert_type)
        if alert:
            control_related = mention.get_mention_control_related()
            send_comment_email_alert(alert, control_related)


def update_related_action_items_description(description, action_item_id, parent_id):
    query_action_item_id = parent_id if parent_id else action_item_id

    ActionItem.objects.filter(
        Q(pk=query_action_item_id) | Q(parent_action_item_id=query_action_item_id)
    ).update(description=description)


def track_implementation_date(control, new_status):
    if new_status == STATUS['IMPLEMENTED']:
        control.implementation_date = timezone.now()
    else:
        control.implementation_date = None
    control.save()


class CreateControl(graphene.Mutation):
    class Arguments:
        input = CreateControlInput(required=True)

    data = graphene.Field(ControlType)

    @login_required
    @transaction.atomic
    @service_exception('Cannot create control')
    @permission_required('control.add_control')
    @create_revision('Created control')
    def mutate(self, info, input):
        organization_id = info.context.user.organization_id
        name = input.get('name')
        description = input.get('description')
        pillar_id = input.get('pillar_id')

        invalid_input = bool(
            pillar_id is not None
            and (name.strip() == '' or description.strip() == '' or pillar_id < 1)
        )
        if invalid_input:
            raise ServiceException(
                'Control with pillar should have name and description'
            )

        control_exists = Control.objects.filter(
            name=input.get('name'), organization_id=organization_id
        ).exists()

        if control_exists:
            raise ServiceException('Control with that name already exists')

        owners = input.get('owner_emails', [])
        validate_owners_quantity(owners)
        validate_owners_empty_values(owners)

        administrator = get_user_by_email(
            organization_id=organization_id, email=input.get('administrator_email')
        )
        approver = get_user_by_email(
            organization_id=organization_id, email=input.get('approver_email')
        )
        pillar_id = input.get('pillar_id')
        control = input.to_model(
            organization_id=info.context.user.organization_id,
            approver=approver,
            administrator=administrator,
            status=STATUS.get('NOT IMPLEMENTED'),
            owners=owners,
            pillar_id=pillar_id,
        )
        tags = input.get('tag_names')
        tag_control(organization_id, control, tags)
        certification_section_ids = input.get('certification_section_ids')
        if certification_section_ids and len(certification_section_ids):
            control.certification_sections.add(*certification_section_ids)
        set_assignees_to_all_child_action_items(control)

        Control.controls_health(organization_id=organization_id, refresh_cache=True)

        return CreateControl(data=control)


class UpdateControl(graphene.Mutation):
    class Arguments:
        input = UpdateControlInput(required=True)

    data = graphene.Field(ControlType)

    @login_required
    @transaction.atomic
    @service_exception('Cannot update control')
    @permission_required('control.change_control')
    @create_revision('Updated control')
    def mutate(self, info, input):
        id = input.get('id')
        organization_id = info.context.user.organization_id
        current_control = Control.objects.get(id=id, organization_id=organization_id)

        if 'owner_emails' in input.keys():
            action_items_override_option = input.get('action_items_override_option')
            owners = input.get('owner_emails')
            validate_owners_quantity(owners)
            validate_owners_empty_values(owners)
            current_control.owners = owners
            set_assignees_to_all_child_action_items(
                current_control, action_items_override_option
            )

        current_control.approver = get_user_by_email(
            organization_id=organization_id, email=input.get('approver_email')
        )

        current_control.administrator = get_user_by_email(
            organization_id=organization_id, email=input.get('administrator_email')
        )

        current_control.has_new_action_items = input.get('has_new_action_items')

        tags = input.get('tag_names')
        update_tag_control(organization_id, current_control, tags)

        new_status = input.get('status', '').upper()
        prev_status = current_control.status

        can_not_implement = current_control.action_items.filter(
            is_required=True
        ).exclude(status='not_applicable').exclude(status='completed').exists() and (
            current_control.status == 'NOT IMPLEMENTED' and new_status == 'IMPLEMENTED'
        )

        if can_not_implement:
            message = (
                'Can not implement control: '
                f'{current_control.name}'
                ' has pending action items'
            )

            raise ServiceException(message)

        if new_status and not can_transition_status(current_control.status, new_status):
            message = f'Transition not allowed:{current_control.status}->{new_status}'
            raise ServiceException(message)

        control = input.to_model(update=current_control)

        if new_status and new_status != prev_status:
            Control.controls_health(organization_id=organization_id, refresh_cache=True)
            track_implementation_date(control, new_status)

        return UpdateControl(data=control)


class DeleteControl(graphene.Mutation):
    class Arguments:
        input = DeleteInput(required=True)

    success = graphene.Boolean(default_value=True)

    @login_required
    @transaction.atomic
    @service_exception('Cannot delete control')
    @permission_required('control.delete_control')
    def mutate(self, info, input):
        id = input.get('id')
        organization_id = info.context.user.organization_id

        control = Control.objects.filter(id=id, organization_id=organization_id)
        with reversion.create_revision():
            reversion.set_comment('Deleted control')
            reversion.set_user(info.context.user)

            control.delete()

            return DeleteControl()


class DeleteControls(graphene.Mutation):
    class Arguments:
        input = DeleteControlsInput(required=True)

    success = graphene.Boolean(default_value=True)

    @service(
        allowed_backends=[
            {
                'backend': ConciergeAuthenticationBackend.BACKEND,
                'permission': 'control.batch_delete_control',
            },
            {
                'backend': AuthenticationBackend.BACKEND,
                'permission': 'control.batch_delete_control',
            },
        ],
        exception_msg='Cannot batch delete controls',
        revision_name='Delete controls',
    )
    def mutate(self, info, input):
        try:
            organization = get_organization_by_user_type(
                info.context.user, input.organization_id
            )

            controls = Control.objects.filter(
                id__in=input.ids, organization_id=organization.id
            )

            for control in controls:
                control.delete()

            return DeleteControls()
        except Exception:
            return DeleteControls(success=False)


class AddControlEvidence(graphene.Mutation):
    class Arguments:
        input = ControlEvidenceInput(required=True)

    evidence_ids = graphene.List(graphene.String)
    duplicated_ids = graphene.List(graphene.String)

    @laika_service(
        permission='control.add_control_evidence',
        exception_msg=GENERIC_FILES_ERROR_MSG,
        revision_name='Evidence added to control',
    )
    def mutate(self, info, input):
        organization = info.context.user.organization
        control = Control.objects.get(pk=input.id)

        duplicated_ids = validate_documents_already_exist(
            organization, control, input.get('documents', [])
        )

        added_files_ids = upload_control_file(
            organization, input.get('files', []), control
        )
        added_policies_ids = add_control_policy(
            organization, input.get('policies', []), control, input.time_zone
        )
        added_laika_papers_ids = add_control_documents_or_laika_papers(
            organization,
            input.get('documents', []),
            control,
            file_type=constants.LAIKA_PAPER,
        )
        added_documents_ids = add_control_documents_or_laika_papers(
            organization, input.get('documents', []), control, file_type=constants.FILE
        )
        added_other_evidence_ids = add_control_other_evidence(
            organization, input.get('other_evidence', []), control
        )
        added_teams_ids = add_control_teams(
            organization, input.get('teams', []), control, input.time_zone
        )
        added_officer_ids = add_control_officers(
            organization, input.get('officers', []), control, input.time_zone
        )
        added_note_ids = add_note_to_control(
            organization, info.context.user, input.get('laika_paper', dict()), control
        )

        evidence_ids = (
            added_files_ids
            + added_policies_ids
            + added_documents_ids
            + added_other_evidence_ids
            + added_teams_ids
            + added_officer_ids
            + added_laika_papers_ids
            + added_note_ids
        )

        return AddControlEvidence(
            evidence_ids=evidence_ids, duplicated_ids=list(duplicated_ids)
        )


class DeleteControlEvidence(graphene.Mutation):
    class Arguments:
        input = DeleteControlEvidenceInput(required=True)

    deleted = graphene.List(graphene.String)

    @login_required
    @transaction.atomic
    @service_exception('Failed to delete evidence from control')
    @permission_required('control.delete_control_evidence')
    def mutate(self, info, input):
        organization = info.context.user.organization
        control = Control.objects.get(id=input.id, organization=organization)
        with reversion.create_revision():
            reversion.set_comment('Deleted control evidence')
            reversion.set_user(info.context.user)

            evidence_to_delete = []
            all_evidence = json.loads(input.evidence[0])

            for evidence in all_evidence:
                evidence_to_delete.append(evidence['id'])

            delete_evidence(organization, evidence_to_delete, control)

            logger.info(
                f'Control evidence ids {evidence_to_delete} in '
                f'organization {organization} deleted'
            )
            return DeleteControlEvidence()


def validate_action_item(ai_input, action_item):
    if 'description' in ai_input:
        if not ai_input.description:
            raise ValidationError('Action item description cannot be empty')
        if not action_item.metadata.get(MetadataFields.IS_CUSTOM.value):
            raise ValidationError('Cannot edit a Non-Custom Action Item.')


def handle_dates_deletion(input, action_item):
    if 'due_date' in input and not input.due_date:
        action_item.due_date = None

    if 'completion_date' in input and not input.completion_date:
        action_item.completion_date = None


class UpdateControls(graphene.Mutation):
    class Arguments:
        input = graphene.List(UpdateControlInput, required=True)

    control_ids = graphene.List(graphene.String)

    @login_required
    @transaction.atomic
    @service_exception('Cannot update controls')
    @permission_required('control.change_control')
    @create_revision('Updated controls')
    def mutate(self, info, input):
        organization = info.context.user.organization
        control_ids = [c.id for c in input]
        controls = Control.objects.filter(id__in=control_ids, organization=organization)

        controls_to_update = []
        for control in controls:
            c = [c for c in input if c.id == control.id][0]

            # I'm just taking into count the status right now because this is
            # the only thing we need to bulk update for controls. In the future
            #  if we would like to bulk update anything else we can change this
            # to update other fields as well
            if c and control.status != c.status:
                control.status = c.status
                controls_to_update.append(control)

        Control.objects.bulk_update(controls_to_update, ['status'])

        return control_ids


class UpdateControlActionItem(graphene.Mutation):
    class Arguments:
        input = UpdateControlActionItemInput(required=True)

    action_item = graphene.Field(ControlActionItemType)

    @laika_service(
        permission='action_item.change_actionitem',
        exception_msg='Failed to update a control action item.',
    )
    def mutate(self, info, input):
        organization_id = info.context.user.organization_id
        action_item = ActionItem.objects.get(pk=input.action_item_id)

        validate_action_item(input, action_item)
        handle_dates_deletion(input, action_item)

        new_status = input.status
        prev_status = action_item.status
        action_item = input.to_model(update=action_item)

        if 'status' in input and prev_status != new_status:
            Control.controls_health(organization_id=organization_id, refresh_cache=True)

            if not (
                new_status == ActionItemStatus.NOT_APPLICABLE
                or prev_status == ActionItemStatus.NOT_APPLICABLE
            ):
                requires_evidence = action_item.metadata.get(
                    MetadataFields.REQUIRED_EVIDENCE.value
                )
                has_evidence = Evidence.objects.filter(
                    action_items__id=input.action_item_id
                )

                is_not_completed = prev_status != ActionItemStatus.COMPLETED
                if (
                    is_truthy(requires_evidence)
                    and is_not_completed
                    and not has_evidence
                ):
                    raise ServiceException('Evidence required')

        if 'owner' in input:
            if input.owner:
                owner = User.objects.get(
                    email=input.owner, organization_id=organization_id
                )
                action_item.assignees.set([owner])
                action_item.create_action_item_alert(
                    sender=info.context.user,
                    receiver=owner,
                    alert_type=ALERT_TYPES['CONTROL_ACTION_ITEM_ASSIGNMENT'],
                    organization_id=organization_id,
                )
            else:
                action_item.assignees.clear()

        action_item.full_clean()
        action_item.save()

        update_access_review_from_action_item(action_item)

        if 'description' in input and action_item.metadata.get(
            MetadataFields.IS_CUSTOM.value
        ):
            update_related_action_items_description(
                input.description,
                action_item.id,
                getattr(action_item.parent_action_item, 'id', None),
            )

        return UpdateControlActionItem(action_item)


class BulkUpdateControlActionItems(graphene.Mutation):
    class Arguments:
        input = UpdateControlActionItemsInput(required=True)

    action_items = graphene.List(ControlActionItemType)

    @laika_service(
        permission='action_item.change_actionitem',
        exception_msg='Failed to update control action items',
    )
    def mutate(self, info, input):
        action_items = annotate_action_items_control_count()
        action_items = action_items.filter(
            controls__id=input.control_id,
            status=ActionItemStatus.NEW,
        ).distinct()

        updated_due_date_ids = bulk_update_control_action_items_due_date(
            action_items, input=input, field='due_date'
        )
        updated_owner_ids = bulk_update_control_action_items_owner(
            action_items, input=input, field='owner'
        )

        updated_action_item_ids = [
            updated_ids
            for ids in [updated_due_date_ids, updated_owner_ids]
            if ids
            for updated_ids in ids
        ]

        return BulkUpdateControlActionItems(
            action_items.filter(id__in=updated_action_item_ids)
        )


class AddControlActionItem(graphene.Mutation):
    class Arguments:
        input = AddControlActionItemInput(required=True)

    action_item = graphene.Field(ControlActionItemType)

    @laika_service(
        permission='action_item.add_actionitem',
        exception_msg='Failed to add a control action item.',
    )
    def mutate(self, info, input):
        organization_id = info.context.user.organization_id
        control = Control.objects.get(id=input.control_id)

        owners = []

        if input.owner:
            owners = User.objects.filter(
                email=input.owner, organization_id=organization_id
            )

        is_recurrent = bool(input.recurrent_schedule)

        pillar = control.pillar
        acronym = pillar.acronym if pillar and pillar.acronym else CUSTOM_PREFIX
        next_reference_id = ActionItem.objects.get_next_index(
            organization=info.context.user.organization,
            prefix=f'{acronym}-C',
        )

        input.metadata[MetadataFields.IS_CUSTOM.value] = True
        input.metadata[MetadataFields.ORGANIZATION_ID.value] = str(organization_id)
        input.metadata[MetadataFields.REFERENCE_ID.value] = next_reference_id
        input.metadata[MetadataFields.TYPE.value] = CONTROL_TYPE
        clean_required_evidence(input)

        action_item = ActionItem.objects.create_shared_action_item(
            name=input.name,
            description=input.description,
            due_date=input.due_date,
            is_required=input.is_required,
            is_recurrent=is_recurrent,
            recurrent_schedule=input.recurrent_schedule,
            metadata=input.metadata,
            users=owners,
        )
        if input.owner:
            owner = User.objects.get(email=input.owner, organization_id=organization_id)
            action_item.assignees.set([owner])
            action_item.create_action_item_alert(
                sender=info.context.user,
                receiver=owner,
                alert_type=ALERT_TYPES['CONTROL_ACTION_ITEM_ASSIGNMENT'],
                organization_id=organization_id,
            )
        control.action_items.add(action_item)
        control.has_new_action_items = True
        control.save()
        return AddControlActionItem(action_item)


def clean_required_evidence(input_value):
    required_evidence = input_value.metadata.get(MetadataFields.REQUIRED_EVIDENCE.value)
    required_evidence = (
        REQUIRED_EVIDENCE_YES if is_truthy(required_evidence) else REQUIRED_EVIDENCE_NO
    )

    input_value.metadata[MetadataFields.REQUIRED_EVIDENCE.value] = required_evidence


def is_truthy(value):
    return value in [True, 'true', 'True', 't', 'T', 'yes', 'Yes', '1', 1]


def migrate(
    subtask: SubTask,
    action_item_ref_id: str,
    organization: Organization,
    assignee: User,
) -> Tuple[List[str], bool]:
    status_detail = []
    mapped = False
    try:
        action_item = ActionItem.objects.get(
            metadata__referenceId=action_item_ref_id,
            metadata__organizationId=str(organization.id),
        )

        if subtask.assignee and not action_item.assignees.exists():
            action_item.assignees.add(subtask.assignee)

        if not subtask.assignee and not action_item.assignees.exists():
            action_item.assignees.add(assignee)

        if subtask.due_date and not action_item.due_date:
            action_item.due_date = subtask.due_date

        if not action_item.completion_date and subtask.completed_on:
            action_item.completion_date = subtask.completed_on

        if subtask.evidence:
            action_item.evidences.add(*subtask.evidence)
            action_item.metadata[
                MetadataFields.REQUIRED_EVIDENCE.value
            ] = CAPITALIZE_YES

        action_item.save()

        subtask.action_item = action_item
        subtask.save()
        mapped = True

    except ActionItem.DoesNotExist as e:
        logger.warning(e)
        status_detail.append(f'Action Item does not exist: {action_item_ref_id}')
        mapped = False
    except Exception as e:
        message = f'A mapping error occurred. Error: {e}'
        logger.warning(message)
        status_detail.append(message)
        mapped = False

    return status_detail, mapped


def migrate_subtasks_to_action_items(organization, assignee):
    status_detail = []
    number_of_mapped_subtasks = 0
    org_subtasks = SubTask.objects.filter(
        task__program__organization_id=organization.id
    )
    mapped_fields = read_mapping_file()
    mapped_fields_keys = mapped_fields.keys()
    for subtask in org_subtasks:
        try:
            if subtask.migration_id not in mapped_fields_keys:
                raise NotSubtaskFoundInMappingFile(
                    f'Subtask with id {subtask.id} '
                    'and migration_id '
                    f'{subtask.migration_id} '
                    'not found in mapping file'
                )

            subtask_detail, mapped = migrate(
                subtask, mapped_fields[subtask.migration_id], organization, assignee
            )
            status_detail.extend(subtask_detail)
            if mapped:
                number_of_mapped_subtasks += 1

        except Exception as error:
            status_detail.append(f'{error}')
    return status_detail, number_of_mapped_subtasks


def get_rows_data(headers, sheet):
    migration_id_action_items_id_pairs = {}
    for row in sheet.iter_rows(min_row=2):
        data = dict(zip(headers, [c.value for c in row[0 : len(headers)]]))

        if not data.get(MIGRATION_ID) or not data.get(LAI_REF_ID):
            logger.warning('There is missing data on columns')
            break

        migration_id_action_items_id_pairs[str(data.get(MIGRATION_ID))] = str(
            data.get(LAI_REF_ID)
        )
    return migration_id_action_items_id_pairs


def read_mapping_file():
    try:
        mapping_file = SeedProfile.objects.get(name=SEED_PROFILE_NAME).file
    except SeedProfile.DoesNotExist:
        raise ServiceException('Missing mapping file in configuration.')

    with ZipFile(mapping_file) as mapping_zip:
        with mapping_zip.open(MAPPING_PROFILE_NAME) as mapping_file:
            logger.info(f'Unzipping the file {MAPPING_PROFILE_NAME}')

            workbook = load_workbook(mapping_file)
            if MAPPING_SHEET not in workbook.sheetnames:
                return ServiceException('Mapping file is invalid.')

            headers = get_headers(workbook[MAPPING_SHEET])

            return get_rows_data(headers, workbook[MAPPING_SHEET])


class MigrateOrganizationToMyCompliance(graphene.Mutation):
    class Arguments:
        payload = MigrateOrganizationPayload(required=True)

    success = graphene.Boolean()

    @concierge_service(
        permission='control.can_migrate_to_my_compliance',
        exception_msg='Failed to migrate organization. Please try again.',
        revision_name='Playbooks to my compliance migration',
    )
    def mutate(self, info, payload):
        try:
            SeedProfile.objects.get(name=SEED_PROFILE_NAME)
        except SeedProfile.DoesNotExist:
            raise ServiceException(
                'We could not find the mapping file in our records 😔'
            )

        frameworks = payload.get('frameworks', [])
        organization = Organization.objects.get(id=payload.get('id'))
        migration = MyComplianceMigration.objects.create(
            organization=organization,
            status=IN_PROGRESS,
            frameworks_detail=', '.join(frameworks),
            mapping_file=SeedProfile.objects.get(name=SEED_PROFILE_NAME).file,
            created_by=info.context.user,
        )

        pool.apply_async(
            migrate_organization,
            args=(
                Organization.objects.get(id=payload.get('id')),
                migration,
                info.context.user,
                payload,
            ),
        )

        return MigrateOrganizationToMyCompliance(success=True)


def migrate_organization(
    organization: Organization, migration, user: User, payload: dict
):
    frameworks = payload.get('frameworks', [])
    assignee = payload.get('assignee', str)
    status_detail = []
    audit_completion_date = None
    try:
        audits = Audit.objects.filter(organization=organization).order_by(
            'completed_at'
        )

        if audits:
            audit_completion_date = audits.first().completed_at

        controls_detail = migrate_controls_to_custom(organization)
        delete_old_playbooks_controls(organization)
        archive_old_certifications(organization)
        frameworks_detail = unlock_frameworks(organization, frameworks)
        seeders_detail = apply_seeders(frameworks, organization, user)
        toggle_feature_flags(organization)
        subtasks_detail, mapped_subtasks = migrate_subtasks_to_action_items(
            organization, User.objects.get(email=assignee)
        )
        implement_controls_and_complete_action_items(
            organization, audit_completion_date, frameworks, assignee
        )

        status_detail.extend(controls_detail)
        status_detail.extend(frameworks_detail)
        status_detail.extend(seeders_detail)
        status_detail.extend(subtasks_detail)

        migration.status = DONE

        all_subtask = SubTask.objects.filter(
            task__program__organization_id=organization.id
        ).count()

        migration.mapped_subtasks = f'{mapped_subtasks}/{all_subtask}'
        logger.info('Migration finished successfully')
    except Exception as e:
        message = f'Error when running organization migration: {e}'
        status_detail.append(message)
        logger.exception(message)
        migration.status = FAILED
    finally:
        migration.status_detail = '\n'.join(status_detail)
        migration.save()
        send_web_socket_alert(user.email, migration.status)


def implement_controls_and_complete_action_items(
    organization: Organization,
    audit_completion_date: Optional[datetime],
    frameworks: list[str],
    assignee: str,
):
    user = get_user(assignee)
    controls = get_controls_by_selected_frameworks(frameworks, organization)

    if controls:
        update_action_items(controls, audit_completion_date, user)
        update_controls(controls, audit_completion_date, user)


def get_controls_by_selected_frameworks(
    frameworks: list[str], organization: Organization
) -> QuerySet[Control]:
    framework_codes = Certification.objects.filter(name__in=frameworks).values_list(
        'code', flat=True
    )
    filter_query = Q(organization=organization)
    code_query = Q()
    for code in framework_codes:
        code_query.add(Q(reference_id__endswith=code), Q.OR)
    filter_query.add(code_query, Q.AND)
    return Control.objects.filter(filter_query)


def get_user(assignee_email: str) -> User:
    return User.objects.get(email=assignee_email)


def update_action_items(
    controls: QuerySet[Control],
    audit_completion_date: Optional[datetime],
    assignee: User,
):
    ActionItem.objects.filter(
        is_required=True, controls__in=controls, status=ActionItemStatus.NEW
    ).update(status=ActionItemStatus.COMPLETED)

    ActionItem.objects.filter(
        is_required=True,
        completion_date__isnull=True,
        controls__in=controls,
    ).update(completion_date=audit_completion_date)

    action_items = ActionItem.objects.filter(
        controls__in=controls, subtasks__isnull=True
    )

    for action_item in action_items:
        action_item.assignees.add(assignee)


def update_controls(
    controls: QuerySet[Control],
    audit_completion_date: Optional[datetime],
    assignee: User,
):
    controls.filter(Q(status=STATUS['NOT IMPLEMENTED']) | Q(status='')).update(
        status=STATUS['IMPLEMENTED'], owner1=assignee
    )
    controls.filter(implementation_date__isnull=True).update(
        implementation_date=audit_completion_date
    )


def apply_seeders(
    frameworks: List[str], organization: Organization, user: User
) -> List[str]:
    status_detail = []

    for framework in frameworks:
        try:
            seed_profile_name = SEED_PROFILES_MAPPING[framework]
            seed_profile = SeedProfile.objects.get(name=seed_profile_name)

            Seed.objects.create(
                organization=organization,
                seed_file=seed_profile.file,
                content_description=seed_profile.content_description,
                created_by=user,
            ).run(run_async=False, should_send_alerts=False)
        except Exception as e:
            message = f'Error seeding the profile {framework}: {e}'
            status_detail.append(message)
            logger.warning(message)
    return status_detail


def toggle_feature_flags(organization: Organization):
    for flag in MY_COMPLIANCE_MIGRATION_FLAGS:
        create_or_enable_flag(organization, flag)

    Flag.objects.filter(name=playbooks_feature_flag, organization=organization).delete()


def archive_old_certifications(organization):
    unlocked_org_certs = UnlockedOrganizationCertification.objects.filter(
        organization=organization
    )
    archived_unlocked_org_certs = [
        ArchivedUnlockedOrganizationCertification(
            certification=unlocked_org_cert.certification, organization=organization
        )
        for unlocked_org_cert in unlocked_org_certs
    ]

    ArchivedUnlockedOrganizationCertification.objects.bulk_create(
        archived_unlocked_org_certs
    )

    unlocked_org_certs.delete()


def unlock_frameworks(organization, frameworks: List[str]) -> List[str]:
    status_detail = []
    for cert_name in frameworks:
        try:
            certification = Certification.objects.get(name=cert_name)
            cert_exists = UnlockedOrganizationCertification.objects.filter(
                organization=organization,
                certification=certification,
            ).exists()

            if not cert_exists:
                UnlockedOrganizationCertification.objects.get_or_create(
                    organization=organization,
                    certification=certification,
                )
        except Exception as e:
            message = f'Error unlocking frameworks: {e}'
            logger.exception(message)
            status_detail.append(message)

    return status_detail


def migrate_controls_to_custom(organization: Organization) -> List[str]:
    return (
        Control.objects.filter(
            Q(organization=organization)
            & (Q(reference_id__isnull=True) | Q(reference_id__contains='CTRL'))
        )
        .exclude(
            (Q(implementation_notes__isnull=True) | Q(implementation_notes=''))
            & Q(evidence__isnull=True)
        )
        .migrate_to_custom()
    )


def delete_old_playbooks_controls(organization: Organization):
    Control.objects.filter(
        Q(organization=organization)
        & (Q(reference_id__isnull=True) | Q(reference_id__contains='CTRL'))
    ).delete()

    logger.info(f'All old controls were deleted. Organization: {organization}')


def send_web_socket_alert(email: str, status: str):
    extra = 'successfully' if status == DONE else 'with some errors'
    send_ws_message_to_group(
        room_id=CX_APP_ROOM,
        sender='Admin',
        receiver=email,
        logger=logger,
        payload=dict(type=status, display_message=f'Migration finished {extra}'),
    )


class UpdateControlFamilyOwner(graphene.Mutation):
    class Arguments:
        input = UpdateControlFamilyOwnerInput(required=True)

    control_family_id = graphene.String()

    @laika_service(
        permission='control.change_control',
        exception_msg='Cannot update control family owner',
        revision_name='Update control family owner',
    )
    def mutate(self, info, input):
        new_action_items_status = [ActionItemStatus.NEW, ActionItemStatus.PENDING]
        control_family_id = input.get('control_family_id')
        owner_email = input.get('owner_email')
        group_id = input.get('group_id')
        organization = info.context.user.organization
        new_owner = User.objects.filter(email=owner_email, organization=organization)

        action_items_to_change_assignee_query = ~Q(assignees__email=owner_email) & Q(
            status__in=new_action_items_status
        )
        controls_to_update_query = Q(pillar__id=control_family_id) & Q(
            organization=organization
        )
        if group_id:
            controls_to_update_query &= Q(group__id=group_id)

        controls_to_update = Control.objects.filter(
            controls_to_update_query
        ).prefetch_related(
            Prefetch(
                'action_items',
                queryset=ActionItem.objects.filter(
                    action_items_to_change_assignee_query
                ).distinct(),
                to_attr='action_items_to_assign',
            )
        )

        controls_to_update_owner = controls_to_update.filter(
            ~Q(owner1__email=owner_email)
        ).distinct()

        for control in controls_to_update:
            for action_item in control.action_items_to_assign:
                action_item.assignees.set(new_owner)

            if control in controls_to_update_owner:
                control.owner1 = new_owner.first()
                control.save()

        return UpdateControlFamilyOwner(control_family_id=control_family_id)

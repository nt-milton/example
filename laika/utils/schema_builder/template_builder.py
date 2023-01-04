import base64
import io
import logging
from dataclasses import dataclass
from typing import Any, Dict, List, Tuple

from django.db.models import Q
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from laika.utils.exceptions import ServiceException
from laika.utils.schema_builder.types.base_field import BaseErrorType, SchemaType
from laika.utils.spreadsheet import CONTENT_TYPE, add_sheet_header
from seeder.seeders.commons import get_formatted_headers

logger = logging.getLogger(__name__)


@dataclass
class SchemaResponse:
    error: Any
    failed_rows: List[Any]
    success_rows: List[Any]


class TemplateBuilder:
    def __init__(
        self,
        schemas: List[SchemaType],
    ):
        self.schemas = schemas
        self.start_row_index = 1

    def _reset_start_row(self):
        self.start_row_index = 1

    def _update_start_row(self):
        self.start_row_index += 1

    def _build_cell(self, field, sheet, header, start_row, column_index: int):
        cell = sheet.cell(row=start_row, column=column_index)
        cell.value = header
        cell.alignment = Alignment(horizontal='center')
        column = get_column_letter(column_index)
        width = max(len(header), (field.get_min_width() / 10))
        sheet.column_dimensions[column].width = width
        return column

    def _build_sheet_instructions(self, schema: SchemaType, sheet):
        for index, field in enumerate(schema.fields):
            instructions = field.instructions if field.instructions else ''
            column_index = index + 1
            self._build_cell(
                field, sheet, instructions, self.start_row_index, column_index
            )
        self._update_start_row()

    def _build_sheet_headers(self, schema: SchemaType, workbook, sheet):
        for index, field in enumerate(schema.fields):
            header = field.get_export_header()
            column_index = index + 1
            column = self._build_cell(
                field, sheet, header, self.start_row_index, column_index
            )
            field.add_validation(
                workbook, sheet, column, f'{schema.header_title} - {header}'
            )
        self._update_start_row()

    def _build_sheets(self, workbook):
        sheet = workbook.active
        for index, schema in enumerate(self.schemas):
            self._reset_start_row()
            sheet = sheet if index == 0 else workbook.create_sheet()
            sheet.title = schema.sheet_name
            if schema.header_title:
                add_sheet_header(len(schema.fields), schema.header_title, sheet)
                self._update_start_row()
            self._build_sheet_headers(schema, workbook, sheet)
            if schema.is_displaying_instructions:
                self._build_sheet_instructions(schema, sheet)

    def _is_valid_excel(self, excel_file):
        if not excel_file.file_name.endswith('.xlsx'):
            raise ServiceException('Invalid file type. File must be .xlsx')

    def _workbook_from_file(self, input_file):
        return load_workbook(io.BytesIO(base64.b64decode(input_file.file)))

    def _validate_headers(self, schema: SchemaType, sheet):
        headers = [cell.value for cell in sheet[self.start_row_index] if cell.value]
        self._update_start_row()

        for field in schema.fields:
            if field.get_export_header() not in headers:
                return ServiceException(f'Missing header {field.get_export_header()}')

    def _validate_sheet(self, schema: SchemaType, workbook):
        if schema.header_title:
            self._update_start_row()
        sheet_names = workbook.sheetnames
        if schema.sheet_name not in sheet_names:
            return ServiceException(f'Missing sheet {schema.sheet_name}')
        return self._validate_headers(schema, workbook[schema.sheet_name])

    def _map_row_to_dict(self, row, headers):
        return dict(zip(headers, [entry for entry in row]))

    def _is_empty_row(self, row, headers):
        return all(not row.get(column).value for column in headers)

    def _validate_row(self, schema: SchemaType, row_dic) -> Any:
        errors = []
        for field in schema.fields:
            value = row_dic[field.name].value if field.name in row_dic else None
            error = field.validate(value)
            if error:
                error.address = row_dic[field.name].coordinate
                errors.append(error)
        return errors

    def _read_rows(self, schema: SchemaType, workbook) -> Tuple[Any, Any]:
        if schema.is_displaying_instructions:
            self._update_start_row()
        sheet = workbook[schema.sheet_name]
        headers = [field.name for field in schema.fields]
        success_rows = []
        failed_rows = []
        for row in sheet.iter_rows(min_row=self.start_row_index):
            row_values = [cell for cell in row]
            row_dic = self._map_row_to_dict(row_values, get_formatted_headers(headers))
            if self._is_empty_row(row_dic, headers):
                continue
            errors = self._validate_row(schema, row_dic)
            if errors:
                failed_rows.extend(errors)
            else:
                success_rows.append(
                    {
                        key: self._get_field_by_name(schema, key).format(
                            row_dic[key].value
                        )
                        for key in row_dic
                    }
                )
        return success_rows, failed_rows

    def _get_field_by_name(self, schema: SchemaType, field_name):
        for field in schema.fields:
            if field.name == field_name:
                return field

    def build(self) -> Any:
        workbook = Workbook()
        self._build_sheets(workbook)
        return workbook

    def build_response(self, filename):
        workbook = self.build()
        response = HttpResponse(
            content=save_virtual_workbook(workbook), content_type=CONTENT_TYPE
        )
        response['Access-Control-Expose-Headers'] = 'Template-Filename'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Template-Filename'] = filename
        return response

    def parse(self, input_file):
        self._is_valid_excel(input_file)
        workbook = self._workbook_from_file(input_file)

        result = {}
        for schema in self.schemas:
            self._reset_start_row()
            response = SchemaResponse(
                error=self._validate_sheet(schema, workbook),
                success_rows=[],
                failed_rows=[],
            )
            result[schema.sheet_name] = response
            if response.error:
                continue

            success_rows, failed_rows = self._read_rows(schema, workbook)
            response.failed_rows = failed_rows
            response.success_rows = success_rows

        return result

    def get_query_filter(self, schema: SchemaType, filters):
        """TODO: Generate query filter from incredible filters."""
        filter_query = Q()

        for element in filters:
            field_name = element['field']
            value = element.get('value', '')
            operator = element['operator']

            field = self._get_field_by_name(schema, field_name)
            field_filter_query = field.get_query_filter(
                field=field.path, value=value, operator=operator
            )
            filter_query.add(field_filter_query, Q.AND)

        return filter_query

    def summarize_errors(self, failed_rows: List[BaseErrorType]):
        errors_by_type: Dict[str, List[str]] = {}
        for row in failed_rows:
            errors = errors_by_type.get(row.type, [])
            errors.append(row.address)
            errors_by_type[row.type] = errors
        return [
            {'type': key, 'addresses': errors_by_type[key]} for key in errors_by_type
        ]

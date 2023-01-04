import base64
import io
import re
from tempfile import NamedTemporaryFile
from typing import Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from laika.utils.exceptions import ServiceException

CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

# Last row in a spreadsheet. Used to select all rows in range
ALL_ROWS = '1048576'

EXCEL_EXTENSION = '.xlsx'


def add_sheet_header(length, display_name, sheet):
    if length:
        column = get_column_letter(length)
        sheet.merge_cells(f'A1:{column}1')
    cell = sheet.cell(1, 1)
    cell.value = display_name
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=12, bold=True)


def add_headers(sheet, headers, row_index=2, bold=False):
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row_index, column=index)
        cell.value = header['name']
        cell.alignment = Alignment(horizontal='center')

        column_letter = get_column_letter(index)
        sheet.column_dimensions[column_letter].width = header['width']
        cell.font = Font(bold=bold)


def add_row_values(sheet, headers, map_row, rows, initial_row, extra_filters=None):
    row_number = initial_row
    for row in rows:
        if extra_filters:
            mapped_row = map_row(headers, row, extra_filters)
        else:
            mapped_row = map_row(headers, row)
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_number, column=index)
            cell.value = mapped_row.get(header.get('key'), '')
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        row_number += 1


def add_list_validation(values, sheet, column):
    validation = DataValidation(type="list", formula1=f'"{",".join(values)}"')
    validation.add(f'{column}3:{column}{ALL_ROWS}')
    sheet.add_data_validation(validation)


def add_range_validation(range_formula, sheet, column):
    validation = DataValidation(type="list", formula1=range_formula)
    validation.add(f'{column}3:{column}{ALL_ROWS}')
    sheet.add_data_validation(validation)


def add_validation_legend(header, values, column, sheet, width):
    cell = sheet.cell(1, column)
    cell.value = header
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=11, bold=True)

    column_letter = get_column_letter(column)
    sheet.column_dimensions[column_letter].width = width

    for index, value in enumerate(values, start=2):
        cell = sheet.cell(index, column)
        cell.value = value


def get_formatted_headers(headings):
    headers = []
    for header in headings:
        # removes line break and extra blank spaces
        if header:
            header_formatted = re.sub("\n|\r", ' ', header)
            header_formatted = re.sub(" +", ' ', header_formatted)
            headers.append(header_formatted)
    return headers


def get_headers(sheet, headers_index):
    rows = sheet.iter_rows(min_row=headers_index, max_row=headers_index)
    first_row = next(rows)
    headings = [c.value for c in first_row]
    return get_formatted_headers(headings)


def validate_if_all_columns_are_empty(dictionary, fields) -> bool:
    if are_columns_empty(dictionary, fields):
        return True

    return False


def are_columns_empty(row, columns):
    return all(not row.get(column) for column in columns)


def create_workbook(file, max_rows, sheet_name, template_headers, headers_index=1):
    if not file.file_name.endswith(EXCEL_EXTENSION):
        raise ServiceException('Invalid file type. File must be .xlsx')

    workbook = load_workbook(io.BytesIO(base64.b64decode(file.file)), False)

    rows_number = 0
    headers_in_file = get_headers(workbook[sheet_name], headers_index)

    initial_data_index = headers_index + 1

    sheet_rows = workbook[sheet_name].iter_rows(min_row=initial_data_index)
    for index, row in enumerate(sheet_rows):
        dictionary = dict(
            zip(headers_in_file, [c.value for c in row[0 : len(headers_in_file)]])
        )

        if validate_if_all_columns_are_empty(dictionary, template_headers):
            index_to_delete = initial_data_index + index
            workbook[sheet_name].delete_rows(index_to_delete)
            break

        rows_number += 1

    if rows_number > max_rows:
        raise ServiceException(
            f'Invalid number of rows: {rows_number}. Maximum permitted                 '
            f' {max_rows}'
        )

    return workbook


def valid_header_names(valid_headers, headers):
    return all(item in valid_headers for item in headers)


def get_workbook_rows(workbook, header_rows_count, valid_headers):
    sheet = workbook.active
    headings = [cell.value for cell in sheet[header_rows_count] if cell.value]
    headers = get_formatted_headers(headings)
    if not valid_header_names(valid_headers, headers):
        return False, [], headers

    rows = []
    initial_row = header_rows_count + 1
    for row in sheet.iter_rows(min_row=initial_row):
        row_num = row[0].row
        row_values = [cell.value for cell in row]
        rows.append((row_num, row_values))
    return True, rows, headers


def add_workbook_sheet(
    workbook: Workbook, sheet_title: str, columns_header: List[Dict[str, str]]
) -> Worksheet:
    sheet = workbook.create_sheet(sheet_title)
    add_headers(sheet, columns_header, 1, bold=True)
    return sheet


def save_virtual_workbook(workbook: Workbook) -> bytes:
    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
    return stream

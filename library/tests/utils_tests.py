from pathlib import Path
from typing import List

import pytest
from openpyxl import load_workbook

from library.constants import NO_RESULT, RESULT_FOUND
from library.models import Question
from library.tests.factory import create_question
from library.utils import (
    are_questions_valid,
    get_questions_annotate_for_fetch,
    get_suggestion_questions,
    validate_match_assign_answer_text,
)
from organization.models import Organization
from seeder.seeders.commons import get_formatted_headers
from user.models import User
from user.tests import create_user

TRANSFORMED_QUESTION_TEXT = 'QuestionExample'
QUESTION_TEXT = 'Question Example'


@pytest.fixture()
def user(graphql_organization: Organization) -> User:
    return create_user(graphql_organization, [], 'user+python+test@heylaika.com')


@pytest.fixture
def headers(sheet):
    HEADER_ROW = 2

    headings = [cell.value for cell in sheet[HEADER_ROW] if cell.value]
    headers = get_formatted_headers(headings)

    return headers


@pytest.fixture
def sheet():
    file_path = Path(__file__).parent / 'resources/LibraryTest.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.active

    return sheet


@pytest.fixture
def sheet_with_errors():
    file_path = Path(__file__).parent / 'resources/LibraryTestWithErrors.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.active

    return sheet


@pytest.fixture()
def question(graphql_organization: Organization) -> Question:
    return create_question(
        graphql_organization, question_text=TRANSFORMED_QUESTION_TEXT
    )


@pytest.fixture()
def default_question(graphql_organization: Organization) -> Question:
    return create_question(
        graphql_organization, question_text=TRANSFORMED_QUESTION_TEXT, default=True
    )


@pytest.mark.functional
def test_get_questions_annotate_for_fetch_tabs(
    graphql_organization: Organization, question: Question
):
    q1 = create_question(
        graphql_organization, question_text='Question             Example'
    )

    match = (
        get_questions_annotate_for_fetch()
        .filter(
            question_without_spaces__iexact=TRANSFORMED_QUESTION_TEXT,
            library_entry__organization=graphql_organization,
        )
        .order_by('-library_entry__updated_at')
        .exclude(id=question.id)
        .exclude(library_entry__answer_text__exact='')
        .first()
    )

    assert match.id == q1.id


@pytest.mark.functional
def test_get_questions_annotate_for_fetch_line_breaks(
    graphql_organization: Organization, question: Question
):
    q1 = create_question(
        graphql_organization,
        question_text='''
            Question
            Example
        ''',
    )

    match = (
        get_questions_annotate_for_fetch()
        .filter(
            question_without_spaces__iexact=TRANSFORMED_QUESTION_TEXT,
            library_entry__organization=graphql_organization,
        )
        .order_by('-library_entry__updated_at')
        .exclude(id=question.id)
        .exclude(library_entry__answer_text__exact='')
        .first()
    )

    assert match.id == q1.id


@pytest.mark.functional
def test_validate_match_assign_answer_text(
    graphql_organization: Organization, question: Question
):
    updated_questions: List[Question] = []
    non_updated_questions: List[Question] = []
    q1 = create_question(
        graphql_organization,
        answer_text='',
        question_text='''
            Question
            Example
        ''',
    )
    validate_match_assign_answer_text(
        question, q1, updated_questions, non_updated_questions
    )
    modified_question = Question.objects.get(id=q1.id)
    assert (
        modified_question.library_entry.answer_text
        == question.library_entry.answer_text
    )
    assert len(modified_question.default_question.equivalent_questions.all()) == 1
    assert updated_questions[0].id == q1.id
    assert len(updated_questions) == 1
    assert len(non_updated_questions) == 0
    assert modified_question.fetch_status == RESULT_FOUND


@pytest.mark.functional
def test_validate_non_match_assign_answer_text(
    graphql_organization: Organization, question: Question
):
    updated_questions: List[Question] = []
    non_updated_questions: List[Question] = []
    q1 = create_question(graphql_organization, answer_text='', question_text='HELLOW')
    validate_match_assign_answer_text(
        None, q1, updated_questions, non_updated_questions
    )
    non_modified_question = Question.objects.get(id=q1.id)
    assert (
        non_modified_question.library_entry.answer_text
        != question.library_entry.answer_text
    )
    assert len(non_modified_question.default_question.equivalent_questions.all()) == 0
    assert non_updated_questions[0].id == q1.id
    assert len(updated_questions) == 0
    assert len(non_updated_questions) == 1
    assert non_modified_question.fetch_status == NO_RESULT


@pytest.mark.functional
def test_validate_match_assign_answer_text_should_have_one_equivalent_question(
    graphql_organization: Organization, question: Question, default_question: Question
):
    updated_questions: List[Question] = []
    non_updated_questions: List[Question] = []
    q1 = create_question(
        graphql_organization,
        answer_text='',
        question_text='Question Example',
        default=False,
    )
    default_question.equivalent_questions.add(q1)
    validate_match_assign_answer_text(
        question, q1, updated_questions, non_updated_questions
    )
    modified_question = Question.objects.get(id=q1.id)
    assert len(modified_question.default_question.equivalent_questions.all()) == 1


def multiline_to_singleline(multiline):
    return ' '.join(multiline.split())


@pytest.mark.functional
def test_are_questions_valid_returns_true_aliases_not_in_default(
    graphql_organization: Organization,
):
    create_question(graphql_organization, default=True, question_text=QUESTION_TEXT)
    create_question(graphql_organization, default=False, question_text=QUESTION_TEXT)
    result = are_questions_valid(
        graphql_organization, QUESTION_TEXT, [TRANSFORMED_QUESTION_TEXT]
    )
    assert result is True


@pytest.mark.functional
def test_are_questions_valid_returns_true_default_not_in_alisases(
    graphql_organization: Organization,
):
    create_question(graphql_organization, default=True, question_text=QUESTION_TEXT)
    create_question(graphql_organization, default=False, question_text=QUESTION_TEXT)
    result = are_questions_valid(
        graphql_organization, TRANSFORMED_QUESTION_TEXT, [QUESTION_TEXT]
    )
    assert result is True


@pytest.mark.functional
def test_get_suggestion_questions_error_existing_question_not_exist(
    graphql_organization: Organization, default_question: Question
):
    try:
        get_suggestion_questions(default_question.id, 123, 9, graphql_organization)
    except Exception as error:
        single_line_error = multiline_to_singleline(str(error))
        expected_single_line_error = multiline_to_singleline(
            """
                Either the existing library question or equivalent
                library question does not exist
            """
        )
        assert single_line_error == expected_single_line_error


@pytest.mark.functional
def test_get_suggestion_questions_error_question_suggestion_not_exist(
    graphql_organization: Organization, default_question: Question
):
    try:
        get_suggestion_questions(123, default_question.id, 9, graphql_organization)
    except Exception as error:
        single_line_error = multiline_to_singleline(str(error))
        expected_single_line_error = multiline_to_singleline(
            """
                Either the existing library question or equivalent
                library question does not exist
            """
        )
        assert single_line_error == expected_single_line_error


@pytest.mark.functional
def test_get_suggestion_questions_error_suggestion_does_not_exist(
    graphql_organization: Organization,
):
    question_1 = create_question(
        graphql_organization, default=True, question_text=QUESTION_TEXT
    )
    question_2 = create_question(
        graphql_organization, default=True, question_text=QUESTION_TEXT
    )
    try:
        get_suggestion_questions(
            question_1.id, question_2.id, question_1.id, graphql_organization
        )
    except Exception as error:
        single_line_error = multiline_to_singleline(str(error))
        expected_single_line_error = multiline_to_singleline(
            """
                Existing library question does not have equivalent
                library question as suggestion
            """
        )
        assert single_line_error == expected_single_line_error


@pytest.mark.functional
def test_get_suggestion_questions_error_chosen_id_does_not_match(
    graphql_organization: Organization,
):
    question_1 = create_question(
        graphql_organization, default=True, question_text=QUESTION_TEXT
    )
    question_2 = create_question(
        graphql_organization, default=True, question_text=QUESTION_TEXT
    )
    question_1.equivalent_suggestions.add(question_2)
    try:
        get_suggestion_questions(question_1.id, question_2.id, 9, graphql_organization)
    except Exception as error:
        single_line_error = multiline_to_singleline(str(error))
        expected_single_line_error = multiline_to_singleline(
            """
                Chosen question id does not match any question
                in the suggestion
            """
        )
        assert single_line_error == expected_single_line_error


@pytest.mark.functional
def test_get_suggestion_questions(graphql_organization: Organization):
    question_1 = create_question(
        graphql_organization, default=True, question_text=QUESTION_TEXT
    )
    question_2 = create_question(
        graphql_organization, default=True, question_text=QUESTION_TEXT
    )
    question_1.equivalent_suggestions.add(question_2)
    existing_question, equivalent_suggestion = get_suggestion_questions(
        question_1.id, question_2.id, question_1.id, organization=graphql_organization
    )
    assert existing_question == question_1
    assert equivalent_suggestion == question_2

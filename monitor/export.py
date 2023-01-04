import dataclasses
import json
from io import BytesIO
from tempfile import NamedTemporaryFile
from typing import Any, Dict, Iterable, Optional

from openpyxl import Workbook

from laika.utils.dates import YYYY_MM_DD_HH_MM_SS
from monitor.constants import EMPTY_RESULTS, RETURN_RESULTS
from monitor.helpers import validate_user_monitor_exclusion_event
from monitor.models import (
    Monitor,
    MonitorExclusion,
    MonitorExclusionEventType,
    MonitorHealthCondition,
    MonitorResult,
    OrganizationMonitor,
)
from monitor.result import Result
from monitor.runner import build_unfiltered_query, dry_run

DEFAULT_SHEET = 'Sheet'


def export_xls(
    org_monitor_id: int,
    include_unfiltered: Optional[bool] = None,
    metadata: Optional[bool] = None,
    limit: Optional[int] = None,
):
    return save_virtual_workbook(
        create_workbook(
            OrganizationMonitor.objects.get(id=org_monitor_id),
            include_unfiltered,
            metadata,
            limit,
        )
    )


def save_virtual_workbook(workbook: Workbook):
    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        output = BytesIO(tmp.read())
    return output


def create_workbook(
    org_monitor: OrganizationMonitor,
    include_unfiltered: Optional[bool] = None,
    metadata: Optional[bool] = None,
    limit: Optional[int] = None,
) -> Workbook:
    if not limit:
        limit = 500
    if include_unfiltered is None:
        include_unfiltered = True
    if metadata is None:
        metadata = True
    workbook = Workbook()
    last_result = get_latest_result(org_monitor)
    build_workbook_sheet(
        workbook, last_result['columns'], last_result['data'], 'Last Result'
    )
    handle_unfiltered_data(org_monitor, workbook, include_unfiltered, limit)
    if metadata:
        add_metadata_tab(workbook, org_monitor)
    add_excluded_data_tab(workbook, org_monitor)
    return workbook


def add_excluded_data_tab(workbook: Workbook, org_monitor: OrganizationMonitor):
    exclude_results = MonitorExclusion.objects.filter(organization_monitor=org_monitor)
    if exclude_results.exists():
        header = [
            'ID excluded',
            'Excluded by',
            'User Id',
            'Exclusion date',
            'Is deleted',
            'Record deprecated',
            'Explanation',
            'Latest date updated',
            'Snapshot',
        ]
        data = get_excluded_data(exclude_results)
        build_workbook_sheet(workbook, header, data, 'Excluded Results')


def get_excluded_data(exclude_results: Iterable[MonitorExclusion]) -> list:
    data = []
    for result in exclude_results:
        monitor_exclusion_event = result.last_event
        user = monitor_exclusion_event.user
        user_id = user.id if user else ''
        event = monitor_exclusion_event.event_type
        data.append(
            [
                f'{result.key} = {result.value}',
                validate_user_monitor_exclusion_event(user),
                user_id,
                result.exclusion_date.strftime(YYYY_MM_DD_HH_MM_SS),
                event == MonitorExclusionEventType.DELETED,
                event == MonitorExclusionEventType.DEPRECATED,
                result.justification,
                monitor_exclusion_event.event_date.strftime(YYYY_MM_DD_HH_MM_SS),
                result.snapshot,
            ]
        )
    return data


def add_metadata_tab(workbook: Workbook, org_monitor: OrganizationMonitor):
    monitor = org_monitor.monitor
    header = [
        'Name',
        'Description',
        'Condition',
        'Run Frequency',
        'Tags',
        'Related Controls',
        'Query',
    ]

    data = [
        monitor.name,
        monitor.description,
        user_friendly_health_condition(monitor.health_condition),
        monitor.frequency,
        format_plain_queryset_values(org_monitor.tags.all(), 'name'),
        format_plain_queryset_values(org_monitor.controls.all(), 'name'),
        monitor.query,
    ]
    build_workbook_sheet(workbook, header, [data], 'Metadata')


def build_workbook_sheet(
    workbook: Workbook, header: Iterable, data: Iterable[list], title: str = None
) -> Workbook:
    if DEFAULT_SHEET in workbook.sheetnames:
        del workbook[DEFAULT_SHEET]
    position = len(workbook.sheetnames)
    if title:
        title = title.replace('[', '').replace(']', '')
    sheet = workbook.create_sheet(title, position)
    sheet.append(header)
    for row in data:
        formatted_data = [arr_to_str(record) for record in row]
        sheet.append(formatted_data)
    return workbook


def get_latest_result(org_monitor: OrganizationMonitor) -> Dict[str, Any]:
    if not MonitorResult.objects.filter(organization_monitor=org_monitor).exists():
        return dataclasses.asdict(Result(columns=[], data=[]))
    monitor_result, *_ = MonitorResult.objects.filter(
        organization_monitor=org_monitor
    ).order_by('-id')[:1]
    return monitor_result.result


def handle_unfiltered_data(
    org_monitor: OrganizationMonitor,
    workbook: Workbook,
    include_unfiltered: bool,
    limit: int,
):
    if include_unfiltered:
        query = build_unfiltered_query(org_monitor.monitor.query, limit)
        unfiltered_data = dry_run(
            org_monitor.organization,
            query,
            org_monitor.monitor.validation_query,
            org_monitor.monitor.runner_type,
        ).to_json()

        build_workbook_sheet(
            workbook,
            unfiltered_data['columns'],
            unfiltered_data['data'],
            'Unfiltered Data',
        )


def user_friendly_health_condition(health_condition: str) -> str:
    if health_condition == MonitorHealthCondition.RETURN_RESULTS:
        return RETURN_RESULTS
    else:
        return EMPTY_RESULTS


def format_plain_queryset_values(queryset, field):
    return ', '.join(list(queryset.values_list(field, flat=True)))


def arr_to_str(data):
    return str(data) if type(data) in (list, dict) else data


def export_monitors(monitors: list[Monitor]) -> str:
    fields = get_monitor_fields()
    export_list = []
    for monitor in monitors:
        export_list.append({k: v for k, v in monitor.__dict__.items() if k in fields})

    return json.dumps(export_list)


def get_monitor_fields():
    exclude_fields = ['id', 'organization', 'parent_monitor']
    return [f.name for f in Monitor._meta.fields if f.name not in exclude_fields]


def import_monitors(raw: str) -> None:
    monitors = json.loads(raw)
    fields = get_monitor_fields()
    for monitor in monitors:
        values = {k: v for k, v in monitor.items() if k in fields}
        Monitor.objects.update_or_create(name=monitor['name'], defaults=values)

import pytest
from openpyxl import Workbook

from monitor.export import (
    DEFAULT_SHEET,
    build_workbook_sheet,
    export_monitors,
    get_excluded_data,
    import_monitors,
)
from monitor.models import Monitor, MonitorExclusion
from monitor.mutations import store_create_event
from monitor.tests.factory import create_organization_monitor


@pytest.mark.functional
def test_get_excluded_data(graphql_client):
    user = graphql_client.context.get('user')
    user.first_name = 'Test name'
    user.save()
    exclusion = MonitorExclusion.objects.create(
        organization_monitor=create_organization_monitor(),
        is_active=True,
        key='monitors.monitor_id',
        value='1',
        snapshot=[''],
    )
    store_create_event(exclusion, user)

    first_row, *_ = get_excluded_data([exclusion])
    condition, user_name, _, _, deleted, deprecated, *_ = first_row

    assert condition == 'monitors.monitor_id = 1'
    assert not deleted
    assert not deprecated
    assert user_name == 'Test name ()'


def test_build_workbook_transform_invalid_chars():
    workbook = Workbook()
    headers = ['column_1', 'column_2']
    data = [['row1_1', 'row1_2']]

    build_workbook_sheet(workbook, headers, data, '[AWS] test')

    assert 'AWS test' in workbook.sheetnames


def test_remove_default_worksheet():
    workbook = Workbook()
    headers = ['column_1', 'column_2']
    data = [['row1_1', 'row1_2']]

    build_workbook_sheet(workbook, headers, data, 'My test')

    assert DEFAULT_SHEET not in workbook.sheetnames


def test_export_json_no_exception():
    workbook = Workbook()
    headers = ['column_1', 'column_2']
    data = [['row1_1', {'Key': 'row1_2'}]]

    sheet_name = 'test'
    build_workbook_sheet(workbook, headers, data, sheet_name)
    assert workbook[sheet_name]


@pytest.mark.functional
def test_export_and_import():
    monitor_name = 'test 1'
    monitor_json = export_monitors([Monitor(id=1, name=monitor_name)])
    import_monitors(monitor_json)
    assert Monitor.objects.filter(name=monitor_name).exists()


@pytest.mark.functional
def test_import_invalid_field_is_ignored():
    monitor_name = 'test 1'
    import_monitors(f'[{{"name": "{monitor_name}", "invalid_field": 123  }}]')
    assert Monitor.objects.filter(name=monitor_name).exists()

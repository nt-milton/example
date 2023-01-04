import json
import logging
import re
from typing import Dict, List

from django.http.response import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.writer.excel import save_virtual_workbook

from laika.auth import login_required, permission_required
from laika.utils.dates import now_date
from laika.utils.spreadsheet import ALL_ROWS, CONTENT_TYPE, add_sheet_header
from objects.models import LaikaObjectType
from objects.types import AttributeTypeFactory

logger = logging.getLogger(__name__)

ATTRIBUTE_ORDER = 'sort_index'
OBJECT_TYPE_ID = 'objectTypeId'
FILTERS = 'filters'
ORDER_BY = 'order_by'
ANNOTATE = 'annotate'


def write_header(sheet, attribute_types, row_number=2):
    for index, attribute_type in enumerate(attribute_types.values()):
        header = attribute_type.get_export_header()
        cell = sheet.cell(row=row_number, column=index + 1)
        cell.value = header
        cell.alignment = Alignment(horizontal='center')

        column = get_column_letter(index + 1)
        width = max(len(header), (attribute_type.get_min_width() / 10))
        sheet.column_dimensions[column].width = width

        validation = attribute_type.get_data_validation()
        if validation:
            validation.add(f'{column}3:{column}{ALL_ROWS}')
            sheet.add_data_validation(validation)


# Invalid characters on file openpyxl/cell/cell.py
def _remove_excel_invalid_characters(text: str) -> str:
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    return ILLEGAL_CHARACTERS_RE.sub(r'', text)


def write_entries(
    sheet, laika_object_type: LaikaObjectType, attribute_types: Dict, **kwargs
):
    entries = laika_object_type.elements.values_list('data', flat=True).filter(
        deleted_at=None
    )

    filters = kwargs.get(FILTERS)
    annotate = kwargs.get(ANNOTATE)

    entries = (
        entries.annotate(**annotate).filter(filters) if filters else entries
    ).order_by('created_at')

    for entry in entries:
        row_data: List = []
        try:
            row_data = [
                attribute_type.get_export_value(entry.get(name))
                for name, attribute_type in attribute_types.items()
            ]

            sheet.append(row_data)
        except IllegalCharacterError:
            organization = laika_object_type.organization
            message = (
                f'Error appending row data {str(row_data)} for '
                f'Laika object type {laika_object_type.display_name} '
                f'in organization {organization}'
            )
            logger.warning(message)

            for idx, value in enumerate(row_data):
                if value:
                    row_data[idx] = _remove_excel_invalid_characters(value)

            sheet.append(row_data)


def write_export_response(
    object_type_id: str,
    laika_object_types: List[LaikaObjectType],
    is_template: bool = False,
    **kwargs,
) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    for index, laika_object_type in enumerate(laika_object_types):
        if index > 0:
            sheet = workbook.create_sheet()

        attribute_types = get_attribute_types(laika_object_type)
        sheet.title = laika_object_type.type_name
        if is_template:
            add_sheet_header(
                len(attribute_types), laika_object_type.display_name, sheet
            )
        if object_type_id == laika_object_type.type_name:
            workbook.active = sheet
        header_row = 1 if not is_template else 2
        write_header(sheet, attribute_types, header_row)

        if not is_template:
            write_entries(sheet, laika_object_type, attribute_types, **kwargs)

    return workbook


def write_fetch_lo_file(
    object_type_id: str,
    laika_object_types: List[LaikaObjectType],
    query: str,
):
    workbook = Workbook()
    sheet = workbook.active
    for index, laika_object_type in enumerate(laika_object_types):
        if index > 0:
            sheet = workbook.create_sheet()

        attribute_types = get_attribute_types(laika_object_type)
        sheet.title = laika_object_type.type_name
        add_sheet_header(len(attribute_types), laika_object_type.display_name, sheet)
        if object_type_id == laika_object_type.type_name:
            workbook.active = sheet

        write_header(sheet, attribute_types)
        write_entries(sheet, laika_object_type, attribute_types)

    # Adding a new tab in the workbook that will contains the query executed
    sheet2 = workbook.create_sheet()
    sheet2.title = 'Query'
    workbook.active = sheet2
    sheet2.append([query])
    return save_virtual_workbook(workbook)


def get_attribute_types(laika_object_type: LaikaObjectType) -> Dict:
    attributes = laika_object_type.attributes.all().order_by(ATTRIBUTE_ORDER)
    return {
        attribute.name: AttributeTypeFactory.get_attribute_type(attribute)
        for attribute in attributes
    }


def get_response(filename, workbook):
    response = HttpResponse(
        content=save_virtual_workbook(workbook), content_type=CONTENT_TYPE
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@csrf_exempt
@login_required
@permission_required('objects.bulk_upload_object')
def export_laika_object(request):
    organization = request.user.organization

    try:
        body_obj = json.loads(request.body.decode('utf-8'))
        object_type_id = body_obj.get(OBJECT_TYPE_ID)
        filters = body_obj.get(FILTERS)

        laika_object_type = LaikaObjectType.objects.get(
            organization=organization, id=object_type_id
        )

        filter_query = laika_object_type.get_incredible_filter_query(filters)
        annotate = laika_object_type.get_annotate(filters)

        time_zone = request.GET.get('timezone')
        date = now_date(time_zone, '%Y_%m_%d')
        file_name = (
            f'{organization.name.upper()}_'
            f'{laika_object_type.type_name.upper()}_'
            f'{date}.xlsx'
        )

        kwargs = {
            FILTERS: filter_query,
            ANNOTATE: annotate,
        }

        workbook = write_export_response(object_type_id, [laika_object_type], **kwargs)
        return get_response(file_name, workbook)
    except Exception as e:
        logger.exception(
            f'Error exporting Laika Object for organization: {organization}. Error: {e}'
        )

    return get_response('Laika_Object.xlsx', Workbook())


@csrf_exempt
@login_required
@permission_required('objects.bulk_upload_object')
def export_template(request):
    organization = request.user.organization
    body_obj = json.loads(request.body.decode('utf-8'))
    object_type_id = body_obj.get(OBJECT_TYPE_ID)
    try:
        laika_object_types = LaikaObjectType.objects.filter(
            organization=organization
        ).order_by('display_name')
        file_name = f'{organization.name.upper()}_LAIKA_OBJECT_TEMPLATE.xlsx'

        workbook = write_export_response(
            object_type_id, laika_object_types, is_template=True
        )
        return get_response(file_name, workbook)

    except Exception as e:
        logger.exception(
            'Error exporting Laika Object template for '
            f'organization: {organization}. Error: {e}'
        )

    return get_response('Laika_Object.xlsx', Workbook())

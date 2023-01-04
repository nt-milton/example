import logging
import uuid
from copy import copy
from multiprocessing.pool import ThreadPool
from typing import Optional
from zipfile import ZipFile

from django.core.exceptions import ValidationError
from django.core.files import File
from django.core.validators import FileExtensionValidator
from django.db import models
from django.db.models import QuerySet
from openpyxl import load_workbook

from alert.models import Alert
from audit.models import Audit
from feature.constants import new_controls_feature_flag
from laika.storage import PrivateMediaStorage
from organization.models import Organization
from seeder.constants import (
    ALL_MY_COMPLIANCE_ORGS,
    DONE,
    IN_PROGRESS,
    PENDING,
    SEED_STATUS,
    UPDATING,
)
from seeder.seeders.commons import (
    are_columns_empty,
    are_columns_required_empty,
    get_headers,
)

USER_MODEL = 'user.User'

logger = logging.getLogger(__name__)
pool = ThreadPool()
STATUS = 'status'
ORGANIZATION_ID = 'Organization ID'
NOT_ENOUGH_DATA = 'Not enough data'


def seeder_file_directory_path(instance, filename):
    if instance.organization:
        entity_id = instance.organization.id
    elif instance.audit:
        entity_id = instance.audit.id
    else:
        entity_id = 'custom_orgs'

    return f'{entity_id}/seeder/{filename}'


def seeder_custom_org_file_directory_path(instance, filename):
    return f'custom_orgs/seeder/{filename}'


def seed_profiles_directory_path(instance, filename):
    return f'seed/profiles/{instance.id}/{filename}'


class ProfileType(models.TextChoices):
    NO_TYPE = '', 'No Type'
    PLAYBOOKS = 'playbooks', 'Playbooks'
    MY_COMPLIANCE = 'my_compliance', 'My Compliance'


class SeedProfile(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4)
    created_at = models.DateField(auto_now_add=True)
    updated_at = models.DateField(auto_now=True)
    name = models.CharField(max_length=100)
    is_visible = models.BooleanField(verbose_name='Visible in Polaris', default=False)
    type = models.CharField(
        verbose_name='Program Type',
        max_length=500,
        blank=True,
        choices=ProfileType.choices,
        default=ProfileType.NO_TYPE,
    )
    content_description = models.TextField(default='', max_length=200)
    default_base = models.BooleanField(blank=False, null=False, default=False)
    file = models.FileField(
        storage=PrivateMediaStorage(),
        upload_to=seed_profiles_directory_path,
        max_length=512,
    )

    def __str__(self):
        return self.name

    def save(self, *args, **kwargs):
        if self.default_base:
            profiles = SeedProfile.objects.filter(default_base=True)
            for profile in profiles:
                profile.default_base = False
                profile.save()
        super(SeedProfile, self).save(*args, **kwargs)


class Seed(models.Model):
    created_at = models.DateField(auto_now_add=True)
    updated_at = models.DateField(auto_now=True)
    organization = models.ForeignKey(
        Organization, on_delete=models.SET_NULL, related_name='seedfiles', null=True
    )
    custom_org_list = models.FileField(
        storage=PrivateMediaStorage(),
        upload_to=seeder_custom_org_file_directory_path,
        validators=[FileExtensionValidator(allowed_extensions=['zip'])],
        blank=True,
        max_length=512,
    )
    profile = models.ForeignKey(
        SeedProfile,
        on_delete=models.SET_NULL,
        related_name='profile',
        blank=True,
        null=True,
    )
    seed_file = models.FileField(
        storage=PrivateMediaStorage(),
        upload_to=seeder_file_directory_path,
        blank=True,
        max_length=512,
    )
    status = models.CharField(max_length=1, choices=SEED_STATUS, default=PENDING)
    status_detail = models.TextField(blank=True)
    created_by = models.ForeignKey(
        USER_MODEL,
        related_name='seeds',
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
    )
    audit = models.ForeignKey(
        Audit, on_delete=models.SET_NULL, related_name='seed_files', null=True
    )
    content_description = models.TextField(default='', max_length=200)

    def clean(self):
        if self.audit:
            return
        if not self.organization and not self.custom_org_list:
            raise ValidationError(
                {
                    'organization': ValidationError(NOT_ENOUGH_DATA),
                    'custom_org_list': ValidationError(NOT_ENOUGH_DATA),
                }
            )

        if (self.custom_org_list or self.organization) and not self.seed_file:
            raise ValidationError(
                {
                    'seed_file': ValidationError('You must select a zip file'),
                }
            )

        if self.seed_file and not self.organization and not self.custom_org_list:
            raise ValidationError(
                {
                    'organization': ValidationError(
                        'Select an organization or submit a list'
                    ),
                }
            )

    def run(self, run_async=True, should_send_alerts=True):
        if (self.organization and self.organization.name == ALL_MY_COMPLIANCE_ORGS) or (
            self.custom_org_list and self.custom_org_list.file
        ):
            self.status = DONE

            if self.custom_org_list and self.custom_org_list.file:
                self.organization = None

            self.save()
            return self

        is_seeding_multiple = self.status == UPDATING
        self.status = IN_PROGRESS
        self.save()

        from seeder.tasks import seed_organization

        if not is_seeding_multiple and run_async:
            seed_organization.delay(
                instance_id=self.id,
                is_updating=is_seeding_multiple,
                should_send_alerts=should_send_alerts,
                run_async=run_async,
            )
            return self

        seed_organization(
            instance_id=self.id,
            is_updating=is_seeding_multiple,
            should_send_alerts=should_send_alerts,
        )

        return Seed.objects.get(id=self.id)

    def create_and_run_upsert_seeds(self, custom_org_list=None):
        if custom_org_list:
            seed_orgs, status = get_organizations_from_zip(custom_org_list)
            self.status_detail = '\n'.join(status)

            if not seed_orgs:
                return self

        else:
            seed_orgs = get_all_my_compliance_orgs()

        for organization in seed_orgs:
            Seed.objects.create(
                organization=organization,
                status=UPDATING,
                seed_file=File(
                    name=self.seed_file.name, file=copy(self.seed_file.file)
                ),
                content_description=self.content_description,
                created_by=self.created_by,
            ).run(run_async=False, should_send_alerts=False)

        return self


def get_all_my_compliance_orgs():
    from feature.models import Flag  # avoids circular import

    return Organization.objects.filter(
        id__in=Flag.objects.filter(
            name=new_controls_feature_flag, is_enabled=True
        ).values_list('organization_id', flat=True)
    )


def get_organizations_from_zip(
    custom_org_zip,
) -> tuple[Optional[QuerySet[Organization]], list[str]]:
    status_detail = []
    fields = [ORGANIZATION_ID, 'Organization Name']
    workbook = None
    sheet_name = 'upsert org list'

    try:
        with ZipFile(custom_org_zip) as seed_zip:
            logger.info('Inside with zipfile')
            with seed_zip.open('custom_org_list.xlsx') as spreadsheet:
                logger.info('Inside custom org list file')
                workbook = load_workbook(spreadsheet)
                logger.info(f'Workbook has been loaded {workbook.sheetnames}')
    except Exception as e:
        message = f'Error getting workbook: {e}'
        logger.warning(message)
        status_detail.append(message)
        return None, status_detail

    if (
        not workbook
        or sheet_name not in workbook.sheetnames
        or workbook[sheet_name].cell(row=2, column=1).value is None
    ):
        status_detail.append('Incorrect format')
        return None, status_detail

    organization_ids = []
    headers = get_headers(workbook[sheet_name])
    for row in workbook[sheet_name].iter_rows(min_row=2):
        dictionary = dict(zip(headers, [c.value for c in row[0 : len(headers)]]))

        if row[0:0] is None:
            return None, status_detail

        if are_columns_empty(dictionary, fields):
            continue

        if are_columns_required_empty(dictionary, [ORGANIZATION_ID]):
            status_detail.append(f'Error: {ORGANIZATION_ID} field is empty')
            continue

        organization_ids.append(dictionary[ORGANIZATION_ID])

    return Organization.objects.filter(id__in=organization_ids), status_detail


class SeedAlert(models.Model):
    alert = models.ForeignKey(
        Alert, related_name='seed_alert', on_delete=models.CASCADE
    )
    seed = models.ForeignKey(
        Seed,
        related_name='seeds',
        on_delete=models.CASCADE,
    )


class MyComplianceMigration(models.Model):
    created_at = models.DateField(auto_now_add=True)
    updated_at = models.DateField(auto_now=True)
    organization = models.ForeignKey(
        Organization,
        on_delete=models.CASCADE,
        related_name='my_compliance_migrations',
    )
    frameworks_detail = models.TextField(blank=True)
    mapping_file = models.FileField(
        storage=PrivateMediaStorage(),
        upload_to=seeder_file_directory_path,
        max_length=512,
    )
    status = models.CharField(max_length=1, choices=SEED_STATUS, default=PENDING)
    status_detail = models.TextField(blank=True)
    mapped_subtasks = models.TextField(blank=True)
    created_by = models.ForeignKey(
        USER_MODEL,
        related_name='my_compliance_migrations',
        on_delete=models.CASCADE,
    )

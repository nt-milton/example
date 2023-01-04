import logging
from time import sleep
from typing import Optional
from zipfile import ZipFile

from django.db import transaction
from openpyxl import load_workbook

from alert.constants import SEEDING_FINISH_REMINDER
from laika.celery import app as celery_app
from monitor.models import OrganizationMonitor
from monitor.signals import set_monitor_controls
from organization.models import ACTIVE
from program.utils.alerts import create_alert

from .constants import CX_APP_ROOM, DONE, FAILED
from .models import Seed, SeedAlert
from .seeders import (
    audit_population,
    certification_logos,
    certification_sections,
    checklist,
    control_action_items,
    control_certification_sections,
    control_groups,
    controls,
    documents,
    fieldwork_criteria,
    fieldwork_evidence,
    fieldwork_fetch_logic,
    fieldwork_requirements,
    fieldwork_tests,
    howtoguide,
    libraries,
    object_type_attributes,
    object_types,
    officers,
    organization_certs,
    policies,
    programs,
    subtasks,
    tasks,
    team_members,
    teams,
    templates,
    trainings,
    users,
    vendors,
)
from .seeders.commons import are_columns_empty, get_headers, should_process_sheet
from .seeders.subtasks import (
    SUBTASK_REQUIRED_FIELDS,
    get_value_from_column_or_default,
    reference_id,
)
from .utils import send_seed_email, send_seed_error_email

logger = logging.getLogger('seeder_task')
LOCALHOST = 'localhost'


@celery_app.task(
    autoretry_for=(Exception,),
    retry_kwargs={'max_retries': 2},
    retry_backoff=True,
    name='Seed Profiles to Organization',
)
def seed_profiles_to_organization(
    user_id: str, organization_id: str, profile_ids: list[str]
):
    from organization.models import Organization
    from seeder.models import Seed, SeedProfile
    from user.models import User

    for profile_id in profile_ids:
        try:
            profile = SeedProfile.objects.get(id=profile_id)
            Seed.objects.create(
                organization=Organization.objects.get(id=organization_id),
                profile=profile,
                seed_file=profile.file,
                created_by=User.objects.get(id=user_id),
            ).run(run_async=False, should_send_alerts=True)
        except Exception as e:
            logger.warning(f'Error seeding organization: {e}')
            return {'success': False}
    return {'success': True}


@celery_app.task(name='Seed Audit Fieldwork Data')
def seed_audit_fieldwork(instance):
    try:
        logger.info(f'Processing async seed for audit: {instance.audit.id}')
        status_detail = []

        with ZipFile(instance.seed_file) as seed_zip:
            logger.info('Inside of zipfile')
            with seed_zip.open('fieldwork_seed.xlsx') as seed_file:
                logger.info('Unzipping the file')
                workbook = load_workbook(seed_file)
                logger.info('workbook has been loaded')

                with transaction.atomic():
                    requirements = fieldwork_requirements.FieldworkRequirement(
                        audit=instance.audit, workbook=workbook
                    )
                    status_detail.extend(requirements.seed())
                    fetch_logic = fieldwork_fetch_logic.FieldworkFetchLogic(
                        audit=instance.audit, workbook=workbook
                    )
                    status_detail.extend(fetch_logic.seed())
                    evidence = fieldwork_evidence.FieldworkEvidence(
                        audit=instance.audit, workbook=workbook
                    )
                    status_detail.extend(evidence.seed())
                    criteria = fieldwork_criteria.FieldworkCriteria(
                        audit=instance.audit, workbook=workbook
                    )
                    status_detail.extend(criteria.seed())
                    tests = fieldwork_tests.FieldworkTests(
                        audit=instance.audit, workbook=workbook
                    )
                    status_detail.extend(tests.seed())
                    populations = audit_population.Population(
                        audit=instance.audit, workbook=workbook, zip_obj=seed_zip
                    )
                    status_detail.extend(populations.seed())

        instance.status = DONE
        instance.status_detail = '\n'.join(status_detail)
    except Exception as e:
        logger.exception('Error processing audit seed file', e)
        instance.status = FAILED
        instance.status_detail = str(e)
    finally:
        instance.save()


@celery_app.task(
    autoretry_for=(Exception,),
    retry_kwargs={'max_retries': 3},
    retry_backoff=True,
    name='Seed Organization',
)
def seed_organization(
    instance_id, is_updating=False, should_send_alerts=True, run_async=False
) -> dict:
    logger.info(f'Starting celery task for seed {instance_id}')

    if run_async:
        sleep(5)

    instance = None
    try:
        instance = Seed.objects.get(id=instance_id)
        logger.info(
            f'Processing async seed for organization: {instance.organization.id}'
        )
        status_detail = []

        with ZipFile(instance.seed_file) as seed_zip:
            logger.info('Inside with zipfile')
            with seed_zip.open('seed.xlsx') as seed_file:
                logger.info('Unzipping the file')
                workbook = load_workbook(seed_file)
                logger.info(f'workbook has been loaded {workbook.sheetnames}')
                validate_subtask(workbook)
                my_compliance = is_my_compliance(workbook)

                status_detail.extend(
                    certification_sections.seed(workbook, is_my_compliance)
                )

                status_detail.extend(
                    organization_certs.OrganizationCertification(
                        organization=instance.organization,
                        workbook=workbook,
                        is_my_compliance=is_my_compliance,
                    ).seed()
                )

                status_detail.extend(
                    control_groups.seed(instance.organization, workbook, is_updating)
                )

                status_detail.extend(
                    controls.execute_seed(instance.organization, workbook, is_updating)
                )

                status_detail.extend(
                    control_certification_sections.execute_seed(
                        instance.organization,
                        workbook,
                        is_updating,
                        is_my_compliance=my_compliance,
                    )
                )

                status_detail.extend(
                    control_action_items.seed(
                        instance.organization, workbook, is_updating
                    )
                )

                with transaction.atomic():
                    status_detail.extend(programs.seed(instance.organization, workbook))
                    status_detail.extend(tasks.seed(instance.organization, workbook))
                    status_detail.extend(
                        howtoguide.seed(instance.organization, workbook)
                    )
                    status_detail.extend(subtasks.seed(instance.organization, workbook))
                status_detail.extend(
                    trainings.seed(instance.organization, seed_zip, workbook)
                )
                status_detail.extend(officers.seed(instance.organization, workbook))
                status_detail.extend(users.seed(instance.organization, workbook))
                status_detail.extend(teams.seed(instance.organization, workbook))
                status_detail.extend(team_members.seed(instance.organization, workbook))
                status_detail.extend(
                    policies.seed(
                        instance.organization, seed_zip, workbook, is_updating
                    )
                )
                status_detail.extend(vendors.seed(seed_zip, workbook, is_my_compliance))
                library = libraries.Library(
                    organization=instance.organization, workbook=workbook
                )
                status_detail.extend(library.seed())
                status_detail.extend(
                    certification_logos.seed(seed_zip, workbook, is_my_compliance)
                )
                status_detail.extend(object_types.seed(instance.organization, workbook))
                status_detail.extend(
                    object_type_attributes.seed(instance.organization, workbook)
                )
                status_detail.extend(
                    documents.Documents(
                        organization=instance.organization,
                        zip_obj=seed_zip,
                        workbook=workbook,
                    ).seed()
                )
                status_detail.extend(
                    templates.Templates(
                        organization=instance.organization,
                        zip_obj=seed_zip,
                        workbook=workbook,
                    ).seed()
                )
                status_detail.extend(checklist.seed(instance.organization, workbook))

        instance.status_detail = '\n'.join(status_detail)
        instance.status = DONE
        logger.info(f'Organization {instance.organization_id} Seed Completed')
    except Exception as e:
        logger.exception(f'Error processing seed file. {e}')
        if instance:
            instance.status = FAILED
            instance.status_detail = str(e)
    finally:
        if instance:
            instance.save()
        if should_send_alerts:
            send_alerts(instance)
        reconcile_monitors(instance)

    return {'success': True}


def send_alerts(instance: Optional[Seed]):
    if instance and instance.status == DONE:
        send_seed_email(instance)

        if instance.created_by:
            from integration.slack.implementation import send_alert_to_slack
            from integration.slack.types import SlackAlert

            slack_alert = SlackAlert(
                alert_type=SEEDING_FINISH_REMINDER, receiver=instance.created_by
            )
            send_alert_to_slack(slack_alert)
            create_alert(
                room_id=CX_APP_ROOM,
                receiver=instance.created_by,
                alert_type=SEEDING_FINISH_REMINDER,
                alert_related_model=SeedAlert,
                alert_related_object={'seed': instance},
            )
    elif instance and instance.status == FAILED:
        send_seed_error_email(instance)


def reconcile_monitors(instance: Optional[Seed]):
    if instance and instance.organization.state == ACTIVE:
        organization_monitors = OrganizationMonitor.objects.filter(
            organization=instance.organization
        )
        for om in organization_monitors:
            set_monitor_controls(om)


def build_record(headers, row):
    return dict(zip(headers, [c.value for c in row[0 : len(headers)]]))


def empty_subtask(r):
    return are_columns_empty(r, SUBTASK_REQUIRED_FIELDS)


def validate_subtask(workbook):
    if should_process_sheet(workbook, 'sub-tasks'):
        subtasks_sheet = workbook['sub-tasks']
        headers = get_headers(subtasks_sheet)
        records = (
            build_record(headers, row) for row in subtasks_sheet.iter_rows(min_row=2)
        )
        for record in records:
            if not empty_subtask(record) and not reference_id(record):
                raise ValueError('subtask_reference_id is required')


def is_my_compliance(workbook) -> bool:
    if should_process_sheet(workbook, 'controls'):
        controls_sheet = workbook['controls']
        headers = get_headers(controls_sheet)
        records = (
            build_record(headers, row) for row in controls_sheet.iter_rows(min_row=2)
        )

        for record in records:
            ctrl_reference_id = get_value_from_column_or_default(
                record, 'reference_id', None
            )

            if ctrl_reference_id and 'CTRL' not in ctrl_reference_id:
                return True

    return False

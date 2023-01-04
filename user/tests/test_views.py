import http
import io
from datetime import datetime, timedelta
from unittest.mock import patch
from uuid import uuid4

import pytest
from django.test import Client
from django.utils import timezone
from okta.models import User as OktaUser
from openpyxl import load_workbook

from conftest import JSON_CONTENT_TYPE
from laika.settings import LOGIN_API_KEY
from laika.utils.dates import ISO_8601_FORMAT
from sso.constants import DONE_ENABLED
from sso.models import IdentityProvider, IdentityProviderDomain
from user.constants import MAGIC_LINK_NOT_FOUND, MAGIC_LINK_TOKEN_EXPIRED, ROLE_VIEWER
from user.models import MagicLink, User
from user.tests import create_user
from user.views import (
    email_validator,
    is_password_expired,
    is_sso,
    is_valid_to_change_okta_password,
    send_verification_code_email,
    validate_user_for_otp,
)

OKTA_get_user_by_email = 'laika.okta.api.OktaApi.get_user_by_email'
OKTA_get_one_time_token = 'laika.okta.api.OktaApi.get_one_time_token'
FAKE_ID = 'fake_id'
ONE_HOUR = 3600
CURRENT_TIME = datetime(2021, 1, 1, tzinfo=timezone.utc)

VALID_OKTA_USER = OktaUser(
    dict(
        id=FAKE_ID,
        passwordChanged=(datetime.now() - timedelta(12)).strftime(ISO_8601_FORMAT),
    )
)

NOT_ALLOWED_TO_CHANGE_OKTA_USER = OktaUser(
    dict(
        id=FAKE_ID,
    )
)

INVALID_DATE_OKTA_USER = OktaUser(dict(id=FAKE_ID, passwordChanged='not valid date'))


@pytest.mark.django_db
def test_is_password_expired(graphql_user):
    user = User.objects.filter(first_name='test').first()
    assert not is_password_expired(user)
    assert not is_password_expired(graphql_user)


@pytest.mark.parametrize(
    ('seconds', 'assertion'),
    [(2 * ONE_HOUR, False), (3 * ONE_HOUR, True), (40 * ONE_HOUR, True)],
)
@pytest.mark.django_db
def test_is_valid_to_change_okta_password(seconds, assertion, graphql_user):
    graphql_user.invitation_sent = timezone.now() - timedelta(seconds=seconds)
    graphql_user.save()

    if assertion:
        assert is_valid_to_change_okta_password(graphql_user)
    else:
        assert not is_valid_to_change_okta_password(graphql_user)


@pytest.mark.django_db
def test_is_valid_to_change_okta_password_no_date(graphql_user):
    assert not is_valid_to_change_okta_password(graphql_user)


@pytest.mark.django_db
def test_is_valid_to_change_okta_password_exception(graphql_user):
    graphql_user.invitation_sent = timezone.now()
    assert not is_valid_to_change_okta_password(graphql_user)


@pytest.mark.django_db
@patch(OKTA_get_user_by_email, return_value=VALID_OKTA_USER)
@patch(OKTA_get_one_time_token, return_value='asdf1234')
def test_get_okta_temporary_password_success(
    get_one_time_token_mock,
    get_okta_user_mock,
    graphql_user,
):
    graphql_user.invitation_sent = timezone.now() - timedelta(seconds=3 * ONE_HOUR)
    graphql_user.save()

    generated_code = '123456'
    MagicLink.objects.create(user=graphql_user, temporary_code=generated_code)
    response = execute_get_okta_temporary_password(graphql_user.email, generated_code)
    get_okta_user_mock.assert_called_once()
    get_one_time_token_mock.assert_called_once()

    assert response
    assert response.status_code == http.HTTPStatus.ACCEPTED

    assert response.content.decode("utf-8") == 'asdf1234'

    db_user = User.objects.get(email=graphql_user.email)
    assert not db_user.password_expired
    assert db_user.invitation_sent


@patch(OKTA_get_user_by_email, return_value=VALID_OKTA_USER)
@patch('user.views.send_verification_code_email')
@pytest.mark.django_db
def test_get_one_time_token_success(
    send_verification_code_email_mock,
    get_user_by_email_mock,
    graphql_user,
):
    graphql_user.invitation_sent = timezone.now() - timedelta(seconds=3 * ONE_HOUR)
    graphql_user.save()

    response = execute_view_call(graphql_user.email)

    assert response
    assert response.status_code == http.HTTPStatus.OK

    get_user_by_email_mock.assert_called_once()
    send_verification_code_email_mock.assert_called_once()

    assert not User.objects.get(email=graphql_user.email).password_expired

    magic_link = MagicLink.objects.get(user=graphql_user)
    assert magic_link.temporary_code


@pytest.mark.django_db
@patch(OKTA_get_user_by_email, return_value=NOT_ALLOWED_TO_CHANGE_OKTA_USER)
def test_get_one_time_token_not_allowed(get_user_by_email_mock, graphql_user):
    graphql_user.invitation_sent = timezone.now() - timedelta(seconds=2 * ONE_HOUR)
    graphql_user.save()

    response = execute_view_call(graphql_user.email)

    assert response
    assert response.status_code == http.HTTPStatus.NOT_ACCEPTABLE
    get_user_by_email_mock.assert_called_once()


def test_send_verification_code_email():
    with patch('user.views.send_email') as send_email_mock:
        send_verification_code_email(email='', template_context={})
        send_email_mock.assert_called_once()


@pytest.mark.django_db
def test_get_otp_from_magic_link_success(graphql_user):
    generated_code = '123456'
    response = execute_get_otp_from_magic_link(
        MagicLink.objects.create(user=graphql_user, temporary_code=generated_code).token
    )
    assert response
    assert response.status_code == http.HTTPStatus.OK
    assert response.content.decode('utf-8') == f'{graphql_user.email}:{generated_code}'


@pytest.mark.django_db
def test_get_otp_from_magic_link_not_found(graphql_user):
    response = execute_get_otp_from_magic_link(str(uuid4()))
    assert response
    assert response.status_code == http.HTTPStatus.NOT_FOUND
    assert response.content.decode("utf-8") == MAGIC_LINK_NOT_FOUND


@patch('user.models.get_today', return_value=CURRENT_TIME)
@pytest.mark.django_db
def test_get_otp_from_magic_link_forbidden(mock_timezone, graphql_user):
    generated_code = '123456'
    token = MagicLink.objects.create(
        user=graphql_user, temporary_code=generated_code
    ).token

    MagicLink.objects.filter(token=token).update(
        updated_at=CURRENT_TIME - timedelta(minutes=21)
    )

    response = execute_get_otp_from_magic_link(token)
    mock_timezone.assert_called_once()
    assert response
    assert response.status_code == http.HTTPStatus.FORBIDDEN
    assert response.content.decode("utf-8") == 'Token expired'


def test_get_one_time_token_invalid_email():
    with pytest.raises(Exception) as excinfo:
        Client(
            HTTP_ORIGIN='http://localhost:3000', HTTP_AUTHORIZATION=LOGIN_API_KEY
        ).post('/user/okta_one_time_token/fake.com')

        assert str(excinfo.value) == 'Invalid Request'


def execute_view_call(email: str = 'fake@email.com'):
    return Client(
        HTTP_ORIGIN='http://localhost:3000', HTTP_AUTHORIZATION=LOGIN_API_KEY
    ).post(
        '/user/okta_one_time_token',
        {'email': email.replace('+', ' ')},
        content_type=JSON_CONTENT_TYPE,
    )


def execute_get_okta_temporary_password(
    email: str = 'fake@email.com', otp: str = '123456'
):
    return Client(
        HTTP_ORIGIN='http://localhost:3000', HTTP_AUTHORIZATION=LOGIN_API_KEY
    ).post(
        '/user/get_okta_temporary_password',
        {'email': email.replace('+', ' '), 'otp': otp},
        content_type=JSON_CONTENT_TYPE,
    )


def execute_get_otp_from_magic_link(token: str = str(uuid4())):
    return Client(
        HTTP_ORIGIN='http://localhost:3000', HTTP_AUTHORIZATION=LOGIN_API_KEY
    ).post(
        '/user/get_otp_from_magic_link',
        {'token': token},
        content_type=JSON_CONTENT_TYPE,
    )


def execute_get_magic_link(token: str = str(uuid4())):
    return Client(
        HTTP_ORIGIN='http://localhost:3000', HTTP_AUTHORIZATION=LOGIN_API_KEY
    ).post(
        f'/user/magic_link/{token}',
        content_type=JSON_CONTENT_TYPE,
    )


@pytest.mark.django_db
def test_is_sso(graphql_organization):
    idp = IdentityProvider.objects.create(
        idp_id='1234-oktaid',
        state=DONE_ENABLED,
        name='Okta Idp',
        organization=graphql_organization,
    )
    IdentityProviderDomain.objects.create(idp_id=idp.id, domain='testsso.com')
    assert is_sso('testsso.com')


def test_email_validator():
    assert not email_validator('')
    assert not email_validator('test@heylaika')
    assert email_validator('test@heylaika.com')


def test_validate_user_for_otp_not_valid_email():
    http_response, okta_user, email = validate_user_for_otp('')
    assert (
        http_response
        and http_response.status_code == http.HTTPStatus.BAD_REQUEST
        and http_response.content.decode("utf-8") == 'Invalid Request'
    )
    assert not okta_user
    assert not email


@pytest.mark.django_db
@patch(OKTA_get_user_by_email, return_value=None)
def test_validate_user_for_otp_not_okta_user(get_user_by_email_mock, graphql_user):
    graphql_user.invitation_sent = timezone.now() - timedelta(seconds=2 * ONE_HOUR)
    graphql_user.save()
    http_response, okta_user, email = validate_user_for_otp(graphql_user.email)
    get_user_by_email_mock.assert_called_once()
    assert (
        http_response
        and http_response.status_code == http.HTTPStatus.NOT_FOUND
        and http_response.content.decode("utf-8")
        == f'User {graphql_user.email} not found'
    )
    assert not okta_user
    assert not email


@pytest.mark.django_db
@patch(OKTA_get_user_by_email, return_value=INVALID_DATE_OKTA_USER)
def test_validate_user_for_otp_exception(get_user_by_email_mock, graphql_user):
    error_msg = 'Not allow to change password before three hours'
    graphql_user.invitation_sent = timezone.now() - timedelta(seconds=2 * ONE_HOUR)
    graphql_user.save()

    http_response, okta_user, email = validate_user_for_otp(graphql_user.email)
    get_user_by_email_mock.assert_called_once()
    assert (
        http_response
        and http_response.status_code == http.HTTPStatus.NOT_ACCEPTABLE
        and http_response.content.decode("utf-8") == error_msg
    )
    assert not okta_user
    assert not email


@pytest.mark.django_db
@patch(OKTA_get_user_by_email, return_value=VALID_OKTA_USER)
def test_validate_user_for_otp_success(get_user_by_email_mock, graphql_user):
    graphql_user.invitation_sent = timezone.now() - timedelta(seconds=3 * ONE_HOUR)
    graphql_user.save()
    http_response, okta_user, email = validate_user_for_otp('test@heylaika.com')
    get_user_by_email_mock.assert_called_once()

    assert not http_response
    assert okta_user == VALID_OKTA_USER
    assert email == 'test@heylaika.com'


@pytest.mark.django_db
@patch(OKTA_get_user_by_email, return_value=NOT_ALLOWED_TO_CHANGE_OKTA_USER)
def test_validate_user_for_otp_not_allowed(get_user_by_email_mock, graphql_user):
    error_msg = 'Not allow to change password before three hours'
    graphql_user.invitation_sent = timezone.now() - timedelta(seconds=1 * ONE_HOUR)
    graphql_user.save()
    http_response, okta_user, email = validate_user_for_otp('test@heylaika.com')
    get_user_by_email_mock.assert_called_once()

    assert (
        http_response
        and http_response.status_code == http.HTTPStatus.NOT_ACCEPTABLE
        and http_response.content.decode("utf-8") == error_msg
    )
    assert not okta_user
    assert not email


@pytest.mark.django_db
def test_user_export_template_full_evidence(http_client, graphql_organization):
    user1 = create_user(
        graphql_organization,
        email='jhon1@heylaika.com',
        role=ROLE_VIEWER,
        first_name='john1',
    )
    user2 = create_user(
        graphql_organization,
        email='jhon2@heylaika.com',
        role=ROLE_VIEWER,
        first_name='john2',
    )
    user2.soft_delete_user()
    user2.save()
    response = http_client.get(
        '/user/export/template?model=full&exclude_users=false&is_evidence=true'
    )
    wb = load_workbook(filename=io.BytesIO(response.content))
    source = wb.active
    assert source['C4'].value == user1.email

    wb.active = wb['Deactivated people']
    source = wb.active
    assert source['C3'].value == user2.email

    assert wb.sheetnames[0] == 'Active people'
    assert wb.sheetnames[1] == 'Deactivated people'


@pytest.mark.django_db
def test_get_magic_link_expired(graphql_user):
    generated_code = '123456'
    magic_link = MagicLink.objects.create(
        user=graphql_user, temporary_code=generated_code
    )
    MagicLink.objects.filter(token=magic_link.token).update(
        updated_at=datetime.now() - timedelta(12)
    )
    response = execute_get_magic_link(magic_link.token)
    assert response
    assert response.status_code == http.HTTPStatus.FORBIDDEN
    assert response.content.decode("utf-8") == MAGIC_LINK_TOKEN_EXPIRED
